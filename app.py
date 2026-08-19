import os
import re
import io
import httpx
from flask import Flask, render_template_string, request, jsonify, send_from_directory, session, redirect, url_for
from flask_cors import CORS

app = Flask(__name__)
CORS(app, supports_credentials=True)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", os.urandom(24).hex())
app.config["SESSION_COOKIE_SAMESITE"] = "None"
app.config["SESSION_COOKIE_SECURE"] = True

@app.before_request
def handle_preflight():
    if request.method == "OPTIONS":
        res = jsonify({"status": "ok"})
        origin = request.headers.get("Origin")
        if origin:
            res.headers["Access-Control-Allow-Origin"] = origin
            res.headers["Access-Control-Allow-Credentials"] = "true"
            res.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization, X-Requested-With"
            res.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
        return res

@app.after_request
def add_cors_headers(response):
    origin = request.headers.get("Origin")
    if origin:
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Access-Control-Allow-Credentials"] = "true"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization, X-Requested-With"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    return response

# ============================================================================
# SUPABASE CONFIGURATION
# ============================================================================
SUPABASE_URL = "https://gscxycvoeprxmkzfvnks.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImdzY3h5Y3ZvZXByeG1remZ2bmtzIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzQzNjA0OTAsImV4cCI6MjA4OTkzNjQ5MH0.CVqLTwGfTCdA2EeSu2ayEv3ID4P68STHgm8XM0c-rus"

def _sb_headers(extra=None):
    h = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}", "Content-Type": "application/json"}
    if extra:
        h.update(extra)
    return h

def sb_auth_signup(email, password):
    r = httpx.post(f"{SUPABASE_URL}/auth/v1/signup", json={"email": email, "password": password}, headers=_sb_headers(), timeout=30.0)
    if r.status_code >= 400:
        data = r.json()
        raise Exception(data.get("msg") or data.get("error_description") or data.get("message") or r.text)
    return r.json()

def sb_auth_signin(email, password):
    r = httpx.post(f"{SUPABASE_URL}/auth/v1/token?grant_type=password", json={"email": email, "password": password}, headers=_sb_headers(), timeout=30.0)
    if r.status_code >= 400:
        data = r.json()
        raise Exception(data.get("error_description") or data.get("msg") or data.get("message") or r.text)
    return r.json()

def sb_select(table, columns="*", filters=None, order=None, limit=None):
    params = {"select": columns}
    if order:
        params["order"] = order
    if limit:
        params["limit"] = str(limit)
    if filters:
        params.update(filters)
    r = httpx.get(f"{SUPABASE_URL}/rest/v1/{table}", params=params, headers=_sb_headers(), timeout=30.0)
    r.raise_for_status()
    return r.json()

def sb_insert(table, data):
    r = httpx.post(f"{SUPABASE_URL}/rest/v1/{table}", json=data, headers=_sb_headers({"Prefer": "return=representation"}), timeout=30.0)
    r.raise_for_status()
    return r.json()

def sb_delete(table, filters):
    r = httpx.delete(f"{SUPABASE_URL}/rest/v1/{table}", params=filters, headers=_sb_headers(), timeout=30.0)
    r.raise_for_status()

# ============================================================================
# SAMPLES SEED DATA
# ============================================================================
SAMPLES = [
    {"sample_no": 1001, "article": "SQUARE FLORAL", "product": "WHITE + PRINT", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "104*072", "construction_total": 176, "blend": "100% VISCOSE", "weave": "TWILL", "finish": "SOFT TOUCH", "gsm": 147},
    {"sample_no": 1002, "article": "HBS DOT PRINT", "product": "DYED + PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "110*070", "construction_total": 180, "blend": "100% COTTON", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 113},
    {"sample_no": 1003, "article": "BELLE MEADE", "product": "DYED", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "140*076", "construction_total": 216, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 135},
    {"sample_no": 1004, "article": "GLENDALE", "product": "WHITE + PRINT", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "132*084", "construction_total": 216, "blend": "100% COTTON", "weave": "DOBBY", "finish": "SOFT TOUCH", "gsm": 109},
    {"sample_no": 1005, "article": "NAVY PEACOAT", "product": "DYED + PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "104*072", "construction_total": 176, "blend": "100% COTTON", "weave": "PLAIN", "finish": "PEACH FIN HAND", "gsm": 111},
    {"sample_no": 1006, "article": "MSHR84 CHAVAL", "product": "WHITE + PRINT", "yarn": "CARDED", "count": "40*40", "count_avg": 40, "construction": "104*072", "construction_total": 176, "blend": "100% COTTON", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 111},
    {"sample_no": 1007, "article": "GARY", "product": "DYED + PRINT", "yarn": "COMPACT", "count": "21*21", "count_avg": 21, "construction": "054*048", "construction_total": 102, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 122},
    {"sample_no": 1008, "article": "NPD 44", "product": "DYED", "yarn": "COMPACT", "count": "30*44", "count_avg": 37, "construction": "076*044", "construction_total": 120, "blend": "COTTON:LENIN", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 133},
    {"sample_no": 1009, "article": "24P5 BL", "product": "WHITE + PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "148*106", "construction_total": 254, "blend": "100% COTTON", "weave": "PLAIN", "finish": "SOFT FIN TOUCH", "gsm": 106},
    {"sample_no": 1010, "article": "f323 091", "product": "DYED + PRINT", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "130*084", "construction_total": 214, "blend": "COTTON:LYCRA", "weave": "TWILL", "finish": "SOFT FIN TOUCH", "gsm": 108},
    {"sample_no": 1011, "article": "MAROO", "product": "WHITE + PRINT", "yarn": "COMPACT YARN", "count": "40*40", "count_avg": 40, "construction": "116*080", "construction_total": 196, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "PEACH FIN HAND", "gsm": 123},
    {"sample_no": 1012, "article": "ALMETA", "product": "CHECKS", "yarn": "SLUB", "count": "40*30", "count_avg": 35, "construction": "080*054", "construction_total": 134, "blend": "100% COTTON", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 119},
    {"sample_no": 1013, "article": "LUMBERTON", "product": "CHECKS", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "094*080", "construction_total": 174, "blend": "100% COTTON", "weave": "DOBBY", "finish": "EASY TO IRON+SOFT TOUCH", "gsm": 144},
    {"sample_no": 1014, "article": "A37900DA", "product": "DYED + PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "128*066", "construction_total": 194, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "PEACH FIN HAND", "gsm": 121},
    {"sample_no": 1015, "article": "A37055PA", "product": "DYED + PRINT", "yarn": "COMPACT", "count": "50*60", "count_avg": 55, "construction": "144*104", "construction_total": 248, "blend": "COTTON:MODAL", "weave": "TWILL", "finish": "EASY TO IRON", "gsm": 118},
    {"sample_no": 1016, "article": "SANGARIA BASE", "product": "DYED + PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "132*072", "construction_total": 204, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 123},
    {"sample_no": 1017, "article": "AG-2220", "product": "DYED + PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "116*080", "construction_total": 196, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "PEACH FIN HAND", "gsm": 123},
    {"sample_no": 1018, "article": "61606V", "product": "DYED", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "195*104", "construction_total": 299, "blend": "100% COTTON", "weave": "DOBBY-SATIN", "finish": "COTTON SOFT FIN", "gsm": 123},
    {"sample_no": 1019, "article": "AW24-DBCH", "product": "DYED + PRINT", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "104*072", "construction_total": 176, "blend": "100%VISCOSE", "weave": "TWILL", "finish": "SOFT FIN TOUCH", "gsm": 147},
    {"sample_no": 1020, "article": "AW24-WBST", "product": "PRINT", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "104*072", "construction_total": 176, "blend": "100%VISCOSE", "weave": "TWILL", "finish": "SOFT FIN TOUCH", "gsm": 147},
    {"sample_no": 1021, "article": "CRECK", "product": "CHECKS", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "064*054", "construction_total": 118, "blend": "COTTON:TENCEL", "weave": "TWILL", "finish": "BRUSHED", "gsm": 196},
    {"sample_no": 1022, "article": "BARBOUR", "product": "DYED", "yarn": "TFO", "count": "60*30", "count_avg": 45, "construction": "200*128", "construction_total": 328, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 273},
    {"sample_no": 1023, "article": "MOUNTAIN - E", "product": "DYED", "yarn": "TFO", "count": "20*20", "count_avg": 20, "construction": "064*054", "construction_total": 118, "blend": "100% COTTON", "weave": "TWILL", "finish": "NORMAL SOFT FIN", "gsm": 294},
    {"sample_no": 1024, "article": "A29600PC", "product": "YD+PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "138*096", "construction_total": 234, "blend": "100% COTTON", "weave": "TWILL", "finish": "NORMAL SOFT FIN", "gsm": 97},
    {"sample_no": 1025, "article": "A37864PB", "product": "YD+PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "062*056", "construction_total": 118, "blend": "100% VISCOSE", "weave": "TWILL", "finish": "BRUSHED", "gsm": 147},
    {"sample_no": 1026, "article": "A37342PA", "product": "DYED +PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "180*104", "construction_total": 284, "blend": "COTTON:MODAL", "weave": "SATIN", "finish": "EASY TO IRON+CALENDER", "gsm": 118, "documents": [{"name": "Test Report.pdf", "url": "/test_report_1026.html"}]},
    {"sample_no": 1027, "article": "CV PRI", "product": "DYED +PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "110*076", "construction_total": 186, "blend": "COTTON:VISCOSE", "weave": "TWILL", "finish": "BRUSHED", "gsm": 117},
    {"sample_no": 1028, "article": "BLUE SNOW FLAKES", "product": "PRINT", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "096*060", "construction_total": 156, "blend": "COTTON:VISCOSE", "weave": "TWILL", "finish": "BRUSHED", "gsm": 131},
    {"sample_no": 1029, "article": "A37238RA", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "180*104", "construction_total": 284, "blend": "COTTON:MODAL", "weave": "SATIN", "finish": "EASY TO IRON+CALENDER", "gsm": 118},
    {"sample_no": 1030, "article": "F326PJSH", "product": "STRIPES", "yarn": "SLUB", "count": "20*20", "count_avg": 20, "construction": "072*062", "construction_total": 134, "blend": "100% COTTON", "weave": "DOBBY", "finish": "NORMAL SOFT FIN", "gsm": 167},
    {"sample_no": 1031, "article": "47F003G", "product": "DYED", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "128*080", "construction_total": 208, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 130},
    {"sample_no": 1032, "article": "62068", "product": "WHITE+PRINT", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "144*092", "construction_total": 236, "blend": "100% COTTON", "weave": "DOBBY", "finish": "SOFT FIN TOUCH", "gsm": 119},
    {"sample_no": 1033, "article": "HARLAN", "product": "YD+PRINT", "yarn": "COMPACT", "count": "40*30", "count_avg": 35, "construction": "120*066", "construction_total": 186, "blend": "100% COTTON", "weave": "TWILL", "finish": "SOFT TOUCH", "gsm": 131},
    {"sample_no": 1034, "article": "MS12-375", "product": "DYED", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "194*104", "construction_total": 298, "blend": "100% COTTON", "weave": "DOBBY*SATIN", "finish": "COTTON SOFT FIN", "gsm": 124},
    {"sample_no": 1035, "article": "MFS-15346", "product": "PRINT", "yarn": "COMPACT", "count": "50*60", "count_avg": 55, "construction": "144*104", "construction_total": 248, "blend": "COTTON:MODAL", "weave": "TWILL", "finish": "SOFT FIN TOUCH", "gsm": 116},
    {"sample_no": 1036, "article": "MFS_16730", "product": "CHECKS", "yarn": "COMPACT", "count": "60*20", "count_avg": 40, "construction": "124*064", "construction_total": 188, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 128},
    {"sample_no": 1037, "article": "MIDLAND", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "154*096", "construction_total": 250, "blend": "100%MODAL", "weave": "DOBBY", "finish": "SOFT TOUCH", "gsm": 112},
    {"sample_no": 1038, "article": "A38235PA", "product": "DYED + PRINT", "yarn": "COMPACT*SLUB", "count": "50*30", "count_avg": 40, "construction": "112*066", "construction_total": 178, "blend": "100% COTTON", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 121},
    {"sample_no": 1039, "article": "A5680", "product": "YD+PIGMENT PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "104*088", "construction_total": 192, "blend": "COTTON:TENCEL", "weave": "TWILL", "finish": "SOFT TOUCH", "gsm": 120},
    {"sample_no": 1040, "article": "MOUNTAIN-B", "product": "DYED", "yarn": "TFO", "count": "20*20", "count_avg": 20, "construction": "064*054", "construction_total": 118, "blend": "100% COTTON", "weave": "DOBBY", "finish": "NORMAL SOFT FIN", "gsm": 294},
    {"sample_no": 1041, "article": "BDLN-0010", "product": "WHITE", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "180*110", "construction_total": 290, "blend": "100% COTTON", "weave": "SATIN", "finish": "ANTI MICROBIAL", "gsm": 122},
    {"sample_no": 1042, "article": "CREW", "product": "WHITE+PRINT", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "104*072", "construction_total": 176, "blend": "100%VISCOSE", "weave": "TWILL", "finish": "ANTI MICROBIAL", "gsm": 146},
    {"sample_no": 1043, "article": "BLOSSOM", "product": "WHITE+PRINT", "yarn": "COMPACT", "count": "40*30", "count_avg": 35, "construction": "120*072", "construction_total": 192, "blend": "COTTON:VISCOSE", "weave": "TWILL", "finish": "SOFT TOUCH+ANTI MICROBIAL", "gsm": 134},
    {"sample_no": 1044, "article": "40017252", "product": "WHITE+PRINT", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "084*068", "construction_total": 152, "blend": "TENCEL:LINEN", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 126},
    {"sample_no": 1045, "article": "OLIVE", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "70*70", "count_avg": 70, "construction": "110*092", "construction_total": 202, "blend": "100% COTTON GIZ", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 73},
    {"sample_no": 1046, "article": "A37681PA", "product": "WHITE+PRINT", "yarn": "COMPACT", "count": "80*70", "count_avg": 75, "construction": "154*120", "construction_total": 274, "blend": "100% COTTON GIZ", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 91},
    {"sample_no": 1047, "article": "16027", "product": "CHECKS", "yarn": "TFO", "count": "80*40", "count_avg": 60, "construction": "104*072", "construction_total": 176, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 111},
    {"sample_no": 1048, "article": "F323PX", "product": "WHITE + PRINT", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "130*084", "construction_total": 214, "blend": "COTTON:LYCRA", "weave": "TWILL", "finish": "SOFT FIN TOUCH", "gsm": 108},
    {"sample_no": 1049, "article": "SMIL08182", "product": "CHECKS", "yarn": "SLUB", "count": "30*30", "count_avg": 30, "construction": "084*076", "construction_total": 160, "blend": "100% COTTON", "weave": "TWILL", "finish": "COTTON SOFT FIN", "gsm": 134},
    {"sample_no": 1050, "article": "TS23FMFW", "product": "DYED + PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "180*104", "construction_total": 284, "blend": "100% COTTON", "weave": "SATIN", "finish": "SOFT FIN TOUCH", "gsm": 119},
    {"sample_no": 1051, "article": "STREM", "product": "STRIPES", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "182*108", "construction_total": 290, "blend": "100% COTTON", "weave": "PLAIN", "finish": "SOFT FIN TOUCH", "gsm": 120},
    {"sample_no": 1052, "article": "GWD567933", "product": "WHITE", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "190*098", "construction_total": 288, "blend": "100% COTTON", "weave": "DOBBY", "finish": "EASY TO IRON+CALENDER", "gsm": 120},
    {"sample_no": 1053, "article": "PALATINE", "product": "DYED", "yarn": "SLUB", "count": "40*40", "count_avg": 40, "construction": "070*050", "construction_total": 120, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 128},
    {"sample_no": 1054, "article": "US88P", "product": "DYED+PRINT", "yarn": "TFO", "count": "60*60", "count_avg": 60, "construction": "120*080", "construction_total": 200, "blend": "100% COTTON", "weave": "TWILL", "finish": "PEACH FIN HAND", "gsm": 168},
    {"sample_no": 1055, "article": "60SCMP", "product": "WHITE+PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "180*104", "construction_total": 284, "blend": "100% COTTON", "weave": "SATIN", "finish": "EASY TO IRON+CALENDER", "gsm": 119},
    {"sample_no": 1056, "article": "FLAN", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "112*084", "construction_total": 196, "blend": "100% VISCOSE", "weave": "TWILL", "finish": "SOFT TOUCH", "gsm": 123},
    {"sample_no": 1057, "article": "OLIVE", "product": "DYED", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "164*076", "construction_total": 240, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 120},
    {"sample_no": 1058, "article": "SUPREME", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "128*069", "construction_total": 197, "blend": "100% MODAL", "weave": "SATIN", "finish": "SOFT TOUCH", "gsm": 72},
    {"sample_no": 1059, "article": "NSL C4", "product": "STRIPES", "yarn": "SLUB", "count": "30*20", "count_avg": 25, "construction": "072*054", "construction_total": 126, "blend": "100% COTTON", "weave": "DOBBY", "finish": "SOFT TOUCH", "gsm": 127},
    {"sample_no": 1060, "article": "BRACKEN", "product": "WHITE+PRINT", "yarn": "COMPACT", "count": "50*40", "count_avg": 45, "construction": "124*084", "construction_total": 208, "blend": "COTTON:TENCEL", "weave": "TWILL", "finish": "SOFT TOUCH", "gsm": 115},
    {"sample_no": 1061, "article": "REDOT", "product": "DYED", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "169*084", "construction_total": 253, "blend": "COTTON:LYCRA", "weave": "TWILL", "finish": "NORMAL SOFT FIN", "gsm": 136},
    {"sample_no": 1062, "article": "STAND STONE", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "21*21", "count_avg": 21, "construction": "072*064", "construction_total": 136, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 151},
    {"sample_no": 1063, "article": "ZELIA", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "50*30", "count_avg": 40, "construction": "112*080", "construction_total": 192, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 122},
    {"sample_no": 1064, "article": "CREAM", "product": "STRIPES", "yarn": "TFO", "count": "80*60", "count_avg": 70, "construction": "128*094", "construction_total": 222, "blend": "100% COTTON", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 120},
    {"sample_no": 1065, "article": "178.161", "product": "WHITE+PRINT", "yarn": "COMPACT*SLUB", "count": "30*30", "count_avg": 30, "construction": "067*056", "construction_total": 123, "blend": "COTTON:VISCOSE", "weave": "PLAIN", "finish": "EASY TO IRON", "gsm": 103},
    {"sample_no": 1066, "article": "A35501", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "180*104", "construction_total": 284, "blend": "100% COTTON", "weave": "SATIN", "finish": "SOFT TOUCH", "gsm": 139},
    {"sample_no": 1067, "article": "GINK", "product": "WHITE+PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "104*084", "construction_total": 188, "blend": "100% MODAL", "weave": "PLAIN", "finish": "EASY TO IRON", "gsm": 78},
    {"sample_no": 1068, "article": "LIC M", "product": "STRIPES", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "128*080", "construction_total": 208, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "PEACH FIN HAND", "gsm": 131},
    {"sample_no": 1069, "article": "DALMATION", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "120*080", "construction_total": 200, "blend": "100% COTTON", "weave": "TWILL", "finish": "PEACH FIN HAND", "gsm": 125},
    {"sample_no": 1070, "article": "NSL 3FT", "product": "STRIPES", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "220*111", "construction_total": 331, "blend": "100% COTTON", "weave": "SATIN", "finish": "EASY TO IRON", "gsm": 139},
    {"sample_no": 1071, "article": "LOOX", "product": "PRINT", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "148*072", "construction_total": 220, "blend": "100% COTTON", "weave": "PLAIN", "finish": "ANTI MICROBIAL", "gsm": 110},
    {"sample_no": 1072, "article": "BOURNE", "product": "DYED", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "176*104", "construction_total": 280, "blend": "100% COTTON", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 110},
    {"sample_no": 1073, "article": "2379R33", "product": "DYED", "yarn": "COMPACT", "count": "15*15", "count_avg": 15, "construction": "053*040", "construction_total": 93, "blend": "COTTON:LINEN", "weave": "TWILL", "finish": "SOFT TOUCH", "gsm": 176},
    {"sample_no": 1074, "article": "BLUE WHITE", "product": "STRIPES", "yarn": "CARDED", "count": "20*20", "count_avg": 20, "construction": "046*040", "construction_total": 86, "blend": "LINEN:VISCOSE", "weave": "PLAIN", "finish": "EASY TO IRON", "gsm": 190},
    {"sample_no": 1075, "article": "PURPLE BLUE", "product": "PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "144*076", "construction_total": 220, "blend": "100% TENCEL", "weave": "TWILL", "finish": "SOFT TOUCH", "gsm": 138},
    {"sample_no": 1076, "article": "TF23", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "180*104", "construction_total": 284, "blend": "100% COTTON", "weave": "SATIN", "finish": "NORMAL SOFT FIN", "gsm": 119},
    {"sample_no": 1077, "article": "F325", "product": "STRIPES", "yarn": "SLUB", "count": "30*30", "count_avg": 30, "construction": "084*072", "construction_total": 156, "blend": "100% COTTON", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 131},
    {"sample_no": 1078, "article": "HABRA", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "70*70", "count_avg": 70, "construction": "180*104", "construction_total": 284, "blend": "100% COTTON", "weave": "TWILL", "finish": "SOFT TOUCH", "gsm": 102},
    {"sample_no": 1079, "article": "PETAL", "product": "DYED", "yarn": "CARDED", "count": "40*40", "count_avg": 40, "construction": "066*052", "construction_total": 118, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 150},
    {"sample_no": 1080, "article": "MFS", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "180*204", "construction_total": 384, "blend": "100% COTTON", "weave": "SATIN", "finish": "COTTON SOFT FIN", "gsm": 119},
    {"sample_no": 1081, "article": "NSL", "product": "YD+PRINT", "yarn": "COMPACT", "count": "60*40", "count_avg": 50, "construction": "052*050", "construction_total": 102, "blend": "100% LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 50},
    {"sample_no": 1082, "article": "VINE PINK", "product": "WHITE+PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "180*104", "construction_total": 284, "blend": "100% COTTON", "weave": "SATIN", "finish": "SOFT TOUCH", "gsm": 119},
    {"sample_no": 1083, "article": "MISSOURI", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "120*058", "construction_total": 178, "blend": "COTTON:MODAL", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 122},
    {"sample_no": 1084, "article": "JANS", "product": "DYED+PRINT", "yarn": "SLUB", "count": "45*21", "count_avg": 33, "construction": "067*050", "construction_total": 117, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 134},
    {"sample_no": 1085, "article": "NLC", "product": "WHITE", "yarn": "CARDED", "count": "21*21", "count_avg": 21, "construction": "064*054", "construction_total": 118, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 139},
    {"sample_no": 1086, "article": "97260", "product": "CHECKS", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "084*064", "construction_total": 148, "blend": "COTTON:VISCOSE", "weave": "TWILL", "finish": "SOFT TOUCH", "gsm": 124},
    {"sample_no": 1087, "article": "NSLC4SB", "product": "CHECKS", "yarn": "SLUB", "count": "40*40", "count_avg": 40, "construction": "094*084", "construction_total": 178, "blend": "COTTON:VISCOSE", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 112},
    {"sample_no": 1088, "article": "F322N", "product": "DYED", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "130*084", "construction_total": 214, "blend": "COTTON:LYCRA", "weave": "TWILL", "finish": "COTTON SOFT FIN", "gsm": 107},
    {"sample_no": 1089, "article": "97264", "product": "CHECKS", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "084*064", "construction_total": 148, "blend": "COTTON:VISCOSE", "weave": "TWILL", "finish": "SOFT TOUCH", "gsm": 124},
    {"sample_no": 1090, "article": "FERGUS", "product": "WHITE+PRINT", "yarn": "TFO", "count": "60*60", "count_avg": 60, "construction": "220*128", "construction_total": 348, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 142},
    {"sample_no": 1091, "article": "ARYALI", "product": "PRINT", "yarn": "COMPACT", "count": "70*70", "count_avg": 70, "construction": "120*100", "construction_total": 220, "blend": "100%COTTON", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 79},
    {"sample_no": 1092, "article": "A38456PA", "product": "YD+PRINT", "yarn": "CARDED", "count": "40*40", "count_avg": 40, "construction": "076*060", "construction_total": 136, "blend": "100%COTTON", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 85},
    {"sample_no": 1093, "article": "50684", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "108*084", "construction_total": 192, "blend": "100%VISCOSE", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 121},
    {"sample_no": 1094, "article": "AMARI", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "108*090", "construction_total": 198, "blend": "100%MODAL", "weave": "TWILL", "finish": "SOFT TOUCH", "gsm": 88},
    {"sample_no": 1095, "article": "BLUE WHITE", "product": "STRIPES", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "110*070", "construction_total": 180, "blend": "100%COTTON", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 113},
    {"sample_no": 1096, "article": "O STIPES", "product": "STRIPES", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "128*080", "construction_total": 208, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 131},
    {"sample_no": 1097, "article": "AW24 YDCH", "product": "CHECKS", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "136*088", "construction_total": 224, "blend": "100% COTTON", "weave": "TWILL", "finish": "COTTON SOFT TOUCH", "gsm": 94},
    {"sample_no": 1098, "article": "DARK NAVY", "product": "DYED", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "068*052", "construction_total": 120, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 199},
    {"sample_no": 1099, "article": "SMIL09", "product": "CHECKS", "yarn": "SLUB", "count": "30*30", "count_avg": 30, "construction": "096*074", "construction_total": 170, "blend": "100%COTTON", "weave": "TWILL", "finish": "COTTON SOFT FIN", "gsm": 142},
    {"sample_no": 1100, "article": "121134", "product": "STRIPES", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "106*084", "construction_total": 190, "blend": "COTTON:MODAL", "weave": "PLAIN", "finish": "SOFT FIN TOUCH", "gsm": 119},
    {"sample_no": 1101, "article": "RED STRIP", "product": "STRIPES", "yarn": "SLUB", "count": "30* 30", "count_avg": 30, "construction": "129 * 098", "construction_total": 227, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 185},
    {"sample_no": 1102, "article": "BOX ORANGE", "product": "CHECKS", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "136*101", "construction_total": 237, "blend": "100%COTTON", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 156},
    {"sample_no": 1103, "article": "4S5321", "product": "CHECKS", "yarn": "SLUB", "count": "40*40", "count_avg": 40, "construction": "132*097", "construction_total": 229, "blend": "100%COTTON", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 179},
    {"sample_no": 1104, "article": "DALLE5", "product": "STRIPES", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "149*101", "construction_total": 250, "blend": "100%COTTON", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 108},
    {"sample_no": 1105, "article": "428GS", "product": "CHECKS", "yarn": "COMBED", "count": "40*40", "count_avg": 40, "construction": "119*098", "construction_total": 217, "blend": "100%COTTON", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 148},
    {"sample_no": 1106, "article": "FLORAOCEAN", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "138*98", "construction_total": 236, "blend": "COTTON:VISCOSE", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 112},
    {"sample_no": 1107, "article": "GEOFLO", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "143*102", "construction_total": 245, "blend": "COTTON:MODAL", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 133},
    {"sample_no": 1108, "article": "5DS7891", "product": "CHECKS", "yarn": "COMBED", "count": "40*40", "count_avg": 40, "construction": "119*92", "construction_total": 211, "blend": "100%COTTON", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 148},
    {"sample_no": 1109, "article": "EARTHBLU", "product": "PRINT", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "122*98", "construction_total": 220, "blend": "COTTON:VISCOSE", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 119},
    {"sample_no": 1110, "article": "4SBLUE", "product": "SRIPES", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "139*102", "construction_total": 241, "blend": "COTTON:MODAL", "weave": "PLAIN", "finish": "EASY TO IRON", "gsm": 116},
    {"sample_no": 1111, "article": "DALLE5", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "142*101", "construction_total": 243, "blend": "100% TENCEL", "weave": "TWILL", "finish": "SOFT TOUCH", "gsm": 119},
    {"sample_no": 1112, "article": "PLAYWILL", "product": "WHITE+PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "129*70", "construction_total": 199, "blend": "100% MODAL", "weave": "SATIN", "finish": "SOFT TOUCH", "gsm": 72},
    {"sample_no": 1113, "article": "54SG87", "product": "PRINT", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "139*102", "construction_total": 241, "blend": "COTTON:VISCOSE", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 148},
    {"sample_no": 1114, "article": "LAVEND", "product": "CHECKS", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "142*102", "construction_total": 244, "blend": "100% COTTON", "weave": "PLAIN", "finish": "ANTI MICROBIAL", "gsm": 110},
    {"sample_no": 1115, "article": "MATRIXY", "product": "CHECKS", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "139*102", "construction_total": 241, "blend": "100% COTTON", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 142},
    {"sample_no": 1116, "article": "SEAGREEN", "product": "DYED", "yarn": "COMBED", "count": "60*60", "count_avg": 60, "construction": "129*096", "construction_total": 225, "blend": "100% COTTON", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 139},
    {"sample_no": 1117, "article": "68493SS", "product": "PRINT", "yarn": "SLUB", "count": "60*60", "count_avg": 60, "construction": "129*096", "construction_total": 225, "blend": "100% TENCEL", "weave": "PLAIN", "finish": "ANTI MICROBIAL", "gsm": 128},
    {"sample_no": 1118, "article": "WHI68", "product": "DYED", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "129*096", "construction_total": 225, "blend": "100% COTTON", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 136},
    {"sample_no": 1119, "article": "FLORALGOLD", "product": "WHITE+PRINT", "yarn": "SLUB", "count": "60*60", "count_avg": 60, "construction": "122*099", "construction_total": 221, "blend": "100% MODAL", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 119},
    {"sample_no": 1120, "article": "NT586", "product": "DYED", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "196*128", "construction_total": 324, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 273},
    {"sample_no": 1121, "article": "7GH79", "product": "PRINT", "yarn": "TFO", "count": "30*30", "count_avg": 30, "construction": "68*52", "construction_total": 120, "blend": "100% LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 143},
    {"sample_no": 1122, "article": "ROSELIT", "product": "CHECKS", "yarn": "CARDED", "count": "30*30", "count_avg": 30, "construction": "106*88", "construction_total": 194, "blend": "100% COTTON", "weave": "TWILL", "finish": "COTTON SOFT FIN", "gsm": 139},
    {"sample_no": 1123, "article": "8745GHJ", "product": "DYED+PRINT", "yarn": "COMBED", "count": "40*40", "count_avg": 40, "construction": "98*72", "construction_total": 170, "blend": "100% VISCOSE", "weave": "PLAIN", "finish": "BRUSHED", "gsm": 139},
    {"sample_no": 1124, "article": "BRICKPALE", "product": "STRIPES", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "68*52", "construction_total": 120, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "COOTON SOFT FIN", "gsm": 149},
    {"sample_no": 1125, "article": "73GJ7", "product": "PRINT", "yarn": "CARDED", "count": "50*50", "count_avg": 50, "construction": "106*88", "construction_total": 194, "blend": "100% COTTON", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 139},
    {"sample_no": 1126, "article": "PRINTMIX67", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "121*71", "construction_total": 192, "blend": "COTTON:MODAL", "weave": "SATIN", "finish": "EASY TO IRON", "gsm": 112},
    {"sample_no": 1127, "article": "ROYAL96", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "128*78", "construction_total": 206, "blend": "COTTON:MODAL", "weave": "SATIN", "finish": "EASY TO IRON", "gsm": 112},
    {"sample_no": 1128, "article": "87CV75", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "196*128", "construction_total": 324, "blend": "100% COTTON", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 137},
    {"sample_no": 1129, "article": "FLORASKU", "product": "DYED+PRINT", "yarn": "COMBED", "count": "60*60", "count_avg": 60, "construction": "121*71", "construction_total": 192, "blend": "100% MODAL", "weave": "SATIN", "finish": "ANTI MICROBIAL", "gsm": 79},
    {"sample_no": 1130, "article": "685VFH", "product": "PRINT", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "139*99", "construction_total": 238, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "PEACH FIN HAND", "gsm": 132},
    {"sample_no": 1131, "article": "8475GJH", "product": "PRINT", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "139*99", "construction_total": 238, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "PEACH FIN HAND", "gsm": 132},
    {"sample_no": 1132, "article": "TEAL987", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "126*78", "construction_total": 204, "blend": "COTTON:MODAL", "weave": "SATIN", "finish": "SOFT TOUCH", "gsm": 112},
    {"sample_no": 1133, "article": "TEALDRA56", "product": "DYED+PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "138*92", "construction_total": 230, "blend": "100% COTTON", "weave": "PLAIN", "finish": "ANTI MICROBIAL", "gsm": 145},
    {"sample_no": 1134, "article": "DIAMOND", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "139*98", "construction_total": 237, "blend": "COTTON:TENCEL", "weave": "PLAIN", "finish": "PEACH FIN HAND", "gsm": 139},
    {"sample_no": 1135, "article": "74PRINT", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "121*92", "construction_total": 213, "blend": "100% COTTON", "weave": "PLAIN", "finish": "PEACH FIN HAND", "gsm": 192},
    {"sample_no": 1136, "article": "ANIMAL97", "product": "PRINT", "yarn": "TFO", "count": "60*60", "count_avg": 60, "construction": "172*102", "construction_total": 274, "blend": "100% MODAL", "weave": "SATIN", "finish": "ANTI MICROBIAL", "gsm": 112},
    {"sample_no": 1137, "article": "CONTINH7Y4", "product": "WHITE+PRINT", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "102*069", "construction_total": 171, "blend": "100% VISCOSE", "weave": "TWILL", "finish": "CALENDER", "gsm": 137},
    {"sample_no": 1138, "article": "79GFD", "product": "PRINT", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "147*78", "construction_total": 225, "blend": "100% COTTON", "weave": "PLAIN", "finish": "CALENDER", "gsm": 145},
    {"sample_no": 1139, "article": "65GV", "product": "PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "168*78", "construction_total": 246, "blend": "100% COTTON", "weave": "PLAIN", "finish": "ANTI MICROBIAL", "gsm": 118},
    {"sample_no": 1140, "article": "LIVE TO", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "168*73", "construction_total": 241, "blend": "100% MODAL", "weave": "SATIN", "finish": "SOFT TOUCH", "gsm": 116},
    {"sample_no": 1141, "article": "DAR9826", "product": "PRINT", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "109*78", "construction_total": 187, "blend": "100% VISCOSE", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 149},
    {"sample_no": 1142, "article": "MARTIX9726", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "152*92", "construction_total": 244, "blend": "100% COTTON", "weave": "PLAIN", "finish": "CALENDER", "gsm": 149},
    {"sample_no": 1143, "article": "PYR746", "product": "PRINT", "yarn": "COMPACT", "count": "50*50", "count_avg": 50, "construction": "148*98", "construction_total": 246, "blend": "COTTON:TENCEL", "weave": "PLAIN", "finish": "CALENDER", "gsm": 152},
    {"sample_no": 1144, "article": "FORESTTREE", "product": "PRINT", "yarn": "COMBED", "count": "60*60", "count_avg": 60, "construction": "138*102", "construction_total": 240, "blend": "COTTON:TENCEL", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 157},
    {"sample_no": 1145, "article": "LOVABLE", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "112*084", "construction_total": 196, "blend": "COTTON:VISCOSE", "weave": "PLAIN", "finish": "EASY TO IRON", "gsm": 119},
    {"sample_no": 1146, "article": "RED DREAM", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "172*107", "construction_total": 279, "blend": "100% MODAL", "weave": "SATIN", "finish": "ANTI MICROBIAL", "gsm": 112},
    {"sample_no": 1147, "article": "795GDY", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "108*090", "construction_total": 198, "blend": "100% MODAL", "weave": "TWILL", "finish": "SOFT TOUCH", "gsm": 60},
    {"sample_no": 1148, "article": "ROLGY6", "product": "PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "067*050", "construction_total": 117, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 134},
    {"sample_no": 1149, "article": "CANN86", "product": "PRINT", "yarn": "SLUB", "count": "40*40", "count_avg": 40, "construction": "121*078", "construction_total": 199, "blend": "COTTON:VISCOSE", "weave": "PLAIN", "finish": "BRUSHED", "gsm": 127},
    {"sample_no": 1150, "article": "MATRI874", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "112*72", "construction_total": 184, "blend": "COTTON:VISCOSE", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 126},
    {"sample_no": 1151, "article": "STRIP47834", "product": "STRIPES", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "112*92", "construction_total": 204, "blend": "100% COTTON", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 145},
    {"sample_no": 1152, "article": "HDSB894", "product": "CHECKS", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "149*96", "construction_total": 245, "blend": "100% COTTON", "weave": "TWILL", "finish": "COTTON SOFT FIN", "gsm": 102},
    {"sample_no": 1153, "article": "GREBOXO", "product": "CHECKS", "yarn": "COMBED", "count": "60*60", "count_avg": 60, "construction": "98*82", "construction_total": 180, "blend": "100% COTTON", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 136},
    {"sample_no": 1154, "article": "975HJH", "product": "STRIPES", "yarn": "SLUB", "count": "60*60", "count_avg": 60, "construction": "102*92", "construction_total": 194, "blend": "100% COTTON", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 152},
    {"sample_no": 1155, "article": "96HJ", "product": "STRIPES", "yarn": "SLUB", "count": "60*60", "count_avg": 60, "construction": "102*92", "construction_total": 194, "blend": "100% COTTON", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 152},
    {"sample_no": 1156, "article": "RAIN8649", "product": "STRIPES", "yarn": "SLUB", "count": "60*60", "count_avg": 60, "construction": "102*92", "construction_total": 194, "blend": "100% COTTON", "weave": "PLAIN", "finish": "ANTI MICROBIAL", "gsm": 139},
    {"sample_no": 1157, "article": "BRICK86", "product": "DYED", "yarn": "SLUB", "count": "60*60", "count_avg": 60, "construction": "128*101", "construction_total": 229, "blend": "100% COTTON", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 198},
    {"sample_no": 1158, "article": "DIA864", "product": "PRINT", "yarn": "TFO", "count": "60*60", "count_avg": 60, "construction": "112*98", "construction_total": 210, "blend": "100% COTTON", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 145},
    {"sample_no": 1159, "article": "G74GFGJD", "product": "YD", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "102*88", "construction_total": 190, "blend": "100% COTTON", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 149},
    {"sample_no": 1160, "article": "DOLEU", "product": "CHECKS", "yarn": "SLUB", "count": "60*60", "count_avg": 60, "construction": "98*77", "construction_total": 175, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 132},
    {"sample_no": 1161, "article": "8946GH", "product": "DYED", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "128*98", "construction_total": 226, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 172},
    {"sample_no": 1162, "article": "INDIG43", "product": "PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "98*72", "construction_total": 170, "blend": "100% COTTON", "weave": "PLAIN", "finish": "CALENDER", "gsm": 139},
    {"sample_no": 1163, "article": "MMON", "product": "STRIPES", "yarn": "COMBEB", "count": "40*40", "count_avg": 40, "construction": "121:98", "construction_total": 219, "blend": "100% COTTON", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 132},
    {"sample_no": 1164, "article": "SOLI7463", "product": "DYED", "yarn": "SLUB", "count": "60*60", "count_avg": 60, "construction": "112*72", "construction_total": 184, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 129},
    {"sample_no": 1165, "article": "09HK", "product": "DYED", "yarn": "TFO", "count": "30*30", "count_avg": 30, "construction": "102*79", "construction_total": 181, "blend": "100% LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 128},
    {"sample_no": 1166, "article": "BVUY", "product": "DYED+PRINT", "yarn": "CARDED", "count": "60*60", "count_avg": 60, "construction": "126*98", "construction_total": 224, "blend": "COTTON:MODAL", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 102},
    {"sample_no": 1167, "article": "STKL", "product": "STRIPES", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "102*88", "construction_total": 190, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 109},
    {"sample_no": 1168, "article": "U89457", "product": "CHECKS", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "119*88", "construction_total": 207, "blend": "100% COTTON", "weave": "PLAIN", "finish": "CALENDER", "gsm": 97},
    {"sample_no": 1169, "article": "INDI028", "product": "PRINT", "yarn": "SLUB", "count": "40*40", "count_avg": 40, "construction": "119*98", "construction_total": 217, "blend": "100% COTTON", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 139},
    {"sample_no": 1170, "article": "JEAN87", "product": "CHECKS", "yarn": "COMBED", "count": "50*50", "count_avg": 50, "construction": "129*102", "construction_total": 231, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 132},
    {"sample_no": 1171, "article": "LIVE79", "product": "CHECKS", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "119*82", "construction_total": 201, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 112},
    {"sample_no": 1172, "article": "MENTO", "product": "CHECKS", "yarn": "SLUB", "count": "40*40", "count_avg": 40, "construction": "102*82", "construction_total": 184, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 114},
    {"sample_no": 1173, "article": "9475", "product": "CHECKS", "yarn": "SLUB", "count": "40*40", "count_avg": 40, "construction": "99*78", "construction_total": 177, "blend": "COOTON:LINEN", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 139},
    {"sample_no": 1174, "article": "CHEKBLUE", "product": "CHECKS", "yarn": "SLUB", "count": "40*40", "count_avg": 40, "construction": "102*88", "construction_total": 190, "blend": "100% COTTON", "weave": "PLAIN", "finish": "BRUSHED", "gsm": 139},
    {"sample_no": 1175, "article": "9762", "product": "CHECKS", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "119*78", "construction_total": 197, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 135},
    {"sample_no": 1176, "article": "9767", "product": "CHECKS", "yarn": "SLUB", "count": "30*30", "count_avg": 30, "construction": "102*79", "construction_total": 181, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 132},
    {"sample_no": 1177, "article": "LILY", "product": "WHITE+PRINT", "yarn": "SLUB", "count": "40*40", "count_avg": 40, "construction": "118*99", "construction_total": 217, "blend": "COOTON:LINEN", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 128},
    {"sample_no": 1178, "article": "7555", "product": "CHECKS", "yarn": "SLUB", "count": "30*30", "count_avg": 30, "construction": "121*98", "construction_total": 219, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 128},
    {"sample_no": 1179, "article": "LOCK89", "product": "CHECKS", "yarn": "SLUB", "count": "30*30", "count_avg": 30, "construction": "121*98", "construction_total": 219, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 136},
    {"sample_no": 1180, "article": "BLUE DOT", "product": "PRINT", "yarn": "SLUB", "count": "40*40", "count_avg": 40, "construction": "102*78", "construction_total": 180, "blend": "100% COTTON", "weave": "PLAIN", "finish": "PEACH FIN HAND", "gsm": 121},
    {"sample_no": 1181, "article": "FORMAL97", "product": "CHECKS", "yarn": "SLUB", "count": "60*60", "count_avg": 60, "construction": "112*99", "construction_total": 211, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 102},
    {"sample_no": 1182, "article": "BLUECHECK", "product": "CHECKS", "yarn": "SLUB", "count": "30*30", "count_avg": 30, "construction": "102*78", "construction_total": 180, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "BRUSHED", "gsm": 108},
    {"sample_no": 1183, "article": "CHE9756", "product": "CHECKS", "yarn": "SLUB", "count": "60*60", "count_avg": 60, "construction": "119*88", "construction_total": 207, "blend": "100% COTTON", "weave": "PLAIN", "finish": "NORMAL SOFT FIN", "gsm": 106},
    {"sample_no": 1184, "article": "MASTURD DOT", "product": "PRINT", "yarn": "COMBEB", "count": "40*40", "count_avg": 40, "construction": "102*78", "construction_total": 180, "blend": "100% COTTON", "weave": "TWILL", "finish": "PEACH FIN HAND", "gsm": 121},
    {"sample_no": 1185, "article": "6543", "product": "CHECKS", "yarn": "SLUB", "count": "30*30", "count_avg": 30, "construction": "112*88", "construction_total": 200, "blend": "100% COTTON", "weave": "PLAIN", "finish": "COTTON SOFT FIN", "gsm": 138},
    {"sample_no": 1186, "article": "MATIX86", "product": "PRINT", "yarn": "SLUB", "count": "30*30", "count_avg": 30, "construction": "98*78", "construction_total": 176, "blend": "COTTON:TENCEL", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 129},
    {"sample_no": 1187, "article": "LEVABEU", "product": "PRINT", "yarn": "COMBEB", "count": "30*30", "count_avg": 30, "construction": "102*88", "construction_total": 190, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 129},
    {"sample_no": 1188, "article": "TEALLEV", "product": "PRINT", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "108*98", "construction_total": 206, "blend": "COTTON:MODAL", "weave": "TWILL", "finish": "SOFT TOUCH", "gsm": 111},
    {"sample_no": 1189, "article": "DOBL826", "product": "CHECKS", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "112*99", "construction_total": 211, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 129},
    {"sample_no": 1190, "article": "PEACH87", "product": "CHECKS", "yarn": "CARDED", "count": "60*60", "count_avg": 60, "construction": "102*79", "construction_total": 181, "blend": "COTTON:MODAL", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 112},
    {"sample_no": 1191, "article": "LENN", "product": "CHECKS", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "98*76", "construction_total": 174, "blend": "COTTON:TENCEL", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 102},
    {"sample_no": 1192, "article": "CR827", "product": "CHECKS", "yarn": "SLUB", "count": "40*40", "count_avg": 40, "construction": "102*88", "construction_total": 190, "blend": "COTTON:LYCRA", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 131},
    {"sample_no": 1193, "article": "STUBU", "product": "DYED+PRINT", "yarn": "SLUB", "count": "60*60", "count_avg": 60, "construction": "128*88", "construction_total": 216, "blend": "100% COTTON", "weave": "PLAIN", "finish": "PEACH FIN HAND", "gsm": 128},
    {"sample_no": 1194, "article": "FLOWM", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "119*98", "construction_total": 217, "blend": "COTTON:MODAL", "weave": "PLAIN", "finish": "PEACH FIN HAND", "gsm": 103},
    {"sample_no": 1195, "article": "DAN896", "product": "PRINT", "yarn": "COMPACT", "count": "30*30", "count_avg": 30, "construction": "102*98", "construction_total": 200, "blend": "COTTON:LINEN", "weave": "PLAIN", "finish": "SOFT TOUCH", "gsm": 131},
    {"sample_no": 1196, "article": "HOLIEM", "product": "PRINT", "yarn": "CARDED", "count": "60*60", "count_avg": 60, "construction": "112*98", "construction_total": 210, "blend": "100% COTTON", "weave": "PLAIN", "finish": "PEACH FIN HAND", "gsm": 136},
    {"sample_no": 1197, "article": "WACO", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "112*98", "construction_total": 210, "blend": "100% COTTON", "weave": "PLAIN", "finish": "BRUSHED", "gsm": 137},
    {"sample_no": 1198, "article": "COFLOW", "product": "PRINT", "yarn": "COMPACT", "count": "40*40", "count_avg": 40, "construction": "118*98", "construction_total": 216, "blend": "100% COTTON", "weave": "PLAIN", "finish": "PEACH FIN HAND", "gsm": 128},
    {"sample_no": 1199, "article": "DA975", "product": "PRINT", "yarn": "SLUB", "count": "30*30", "count_avg": 30, "construction": "121*99", "construction_total": 220, "blend": "100% COTTON", "weave": "PLAIN", "finish": "BRUSHED", "gsm": 128},
    {"sample_no": 1200, "article": "MATRI8", "product": "PRINT", "yarn": "COMPACT", "count": "60*60", "count_avg": 60, "construction": "113*77", "construction_total": 190, "blend": "100% COTTON", "weave": "TWILL", "finish": "PEACH FIN HAND", "gsm": 129},
]

def seed_database():
    """Seed the Supabase samples table if empty."""
    try:
        data = sb_select("samples", columns="sample_no", limit=1)
        if not data:
            sb_insert("samples", SAMPLES)
    except Exception:
        pass

def get_all_samples():
    """Get all samples from Supabase as list of dicts."""
    return sb_select("samples", order="sample_no")

# ============================================================================
# FEEL TERMS DICTIONARY
# ============================================================================
FEEL_DICTIONARY = {
    "Soft Feel": ["soft handfeel", "soft touch", "silky touch", "smooth feel", "peach finish", "soft fleece", "soft", "silky", "smooth", "peach"],
    "Good Drape": ["drapey", "drapy", "good fall", "flowy", "fluid", "elegant fall", "drape", "flow", "good drape"],
    "Shiny": ["high shine", "lustrous", "glossy", "sheen", "bright surface", "satin look", "shiny", "shine", "gloss", "lustre"],
    "Crisp": ["crisp", "stiff", "paper touch", "structured feel", "firm hand", "firm", "crisp look"],
    "Stretchable": ["stretchable", "stretch", "elastic", "lycra", "flex", "flexible", "spandex"],
    "Easy Care": ["easy care", "easy iron", "wrinkle free", "wrinkle resistant", "low maintenance", "wrinkle", "easy"],
    "Textured": ["textured", "grainy", "slub look", "uneven surface", "raw texture", "texture", "slub"],
    "Dense": ["durable", "strong", "long life", "sturdy", "dense", "thick", "heavy"],
    "Anti Microbial": ["anti microbial", "antimicrobial", "anti-microbial", "antibacterial", "anti bacterial", "germ resistant", "hygienic"],
}

# ============================================================================
# RULE TABLE
# ============================================================================
RULE_TABLE = {
    "Soft Feel": {
        "yarn": {"values": ["compact", "combed", "slub", "tfo"], "priority": "HIGH"},
        "count": {"min": 30, "priority": "HIGH"},
        "blend": {"values": ["modal", "tencel", "viscose", "giz", "cotton"], "pure_first": True, "priority": "HIGH"},
        "blend_exclude": ["linen", "lenin"],
        "finish": {"values": ["soft touch", "brushed", "peach fin hand", "normal soft fin", "soft fin touch", "cotton soft fin", "chemical soft fin", "calender", "easy to iron", "anti microbial"], "priority": "HIGH"},
        "yarn_exclude": [],
        "negative_cross": [
            {"blend": "100% cotton", "finish": ["anti microbial", "easy to iron"], "finish_override": ["soft touch", "brushed", "peach fin hand", "soft fin touch", "cotton soft fin"]}
        ],
    },
    "Good Drape": {
        "yarn": {"values": ["compact", "tfo", "combed", "slub"], "priority": "HIGH"},
        "blend": {"values": ["viscose", "tencel", "modal", "giz"], "pure_first": True, "priority": "HIGH"},
        "blend_exclude": ["100% cotton", "cotton:lycra", "cotton:linen", "cotton:lenin"],
        "blend_allow": ["giz"],
        "weave": {"values": ["satin", "twill", "dobby", "plain"], "priority": "HIGH"},
        "gsm": {"max": 200, "priority": "HIGH"},
        "yarn_exclude": [],
        "negative_cross": [],
    },
    "Shiny": {
        "yarn": {"values": ["compact", "tfo"], "priority": "HIGH"},
        "count": {"min": 30, "priority": "HIGH"},
        "blend": {"values": ["tencel", "viscose", "modal", "giz", "cotton"], "pure_first": True, "priority": "HIGH"},
        "blend_exclude": ["linen", "lenin", "cotton:lycra"],
        "weave": {"values": ["twill", "satin", "dobby", "plain"], "priority": "HIGH"},
        "finish": {"values": ["calender", "calendar", "easy to iron"], "priority": "LOW"},
        "yarn_exclude": [],
        "negative_cross": [
            {"blend": "100% cotton", "weave": ["plain", "twill", "dobby", "satin"],
             "weave_override": ["dobby-satin", "dobby*satin"],
             "finish_override": ["anti microbial", "calender"]}
        ],
    },
    "Crisp": {
        "blend": {"values": ["linen", "lenin", "cotton"], "priority": "HIGH"},
        "weave": {"values": ["plain"], "priority": "LOW"},
        "yarn_exclude": [],
        "negative_cross": [
            {"blend": "100% cotton", "weave": ["twill", "satin", "dobby", "hbt"]}
        ],
    },
    "Stretchable": {
        "blend": {"values": ["lycra", "modal", "tencel", "viscose"], "priority": "HIGH"},
        "weave": {"values": ["twill", "satin", "dobby", "plain"], "priority": "HIGH"},
        "yarn_exclude": [],
        "negative_cross": [],
    },
    "Easy Care": {
        "finish": {"values": ["easy to iron", "resin"], "priority": "HIGH"},
        "yarn_exclude": [],
        "negative_cross": [],
    },
    "Textured": {
        "weave": {"values": ["dobby", "twill", "plain"], "priority": "HIGH"},
        "yarn_exclude": [],
        "negative_cross": [],
    },
    "Dense": {
        "weave": {"values": ["matt", "twill", "plain", "dobby"], "priority": "HIGH"},
        "gsm": {"min": 190, "priority": "HIGH"},
        "yarn_exclude": [],
        "negative_cross": [],
    },
    "Anti Microbial": {
        "finish": {"values": ["anti microbial"], "priority": "HIGH"},
        "yarn_exclude": [],
        "negative_cross": [],
    },
}

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def normalize_product(product):
    """Normalize product type by removing spaces around + for consistent matching"""
    return re.sub(r'\s*\+\s*', '+', product.upper().strip())

def contains_any(text, patterns):
    """Check if text contains any of the patterns (case-insensitive)"""
    text_lower = text.lower()
    for pattern in patterns:
        if pattern.lower() in text_lower:
            return True
    return False

def get_blend_order_index(sample, standard_terms):
    """Return the lowest blend-values index matched across all active rules.
    Samples matching the first value in the list sort before those matching later values."""
    best = float('inf')
    for term in standard_terms:
        rule = RULE_TABLE.get(term)
        if not rule or "blend" not in rule:
            continue
        blend_values = rule["blend"].get("values", [])
        sample_blend = sample["blend"].lower()
        for idx, val in enumerate(blend_values):
            if val.lower() in sample_blend:
                if idx < best:
                    best = idx
                break
    return best if best != float('inf') else 9999

def calculate_priority_score(sample, standard_terms):
    """Calculate priority ranking score based on BLEND only.
    Follows the exact sequence in the rule table's blend values list.
    Pure 100% fiber ranks above mixed blends of the same fiber.
    Higher score = better match."""
    total_score = 0

    for term in standard_terms:
        rule = RULE_TABLE.get(term)
        if not rule:
            continue

        if "blend" not in rule or not isinstance(rule["blend"], dict):
            continue

        blend_values = rule["blend"].get("values", [])
        sample_blend = sample["blend"].lower().strip()
        max_positions = len(blend_values)

        # Find the highest-priority fiber that matches
        best_fiber_idx = None
        for idx, val in enumerate(blend_values):
            if val.lower() in sample_blend:
                best_fiber_idx = idx
                break

        if best_fiber_idx is not None:
            # Earlier position in list = higher score
            fiber_score = (max_positions - best_fiber_idx) * 1000

            # Pure/100% blend bonus: 100% of a fiber ranks above mixed
            is_pure = ("100%" in sample_blend) and (":" not in sample_blend) and ("/" not in sample_blend)
            if is_pure:
                # If pure_first flag is set, use large bonus so ALL pure blends
                # rank above ALL mixed blends regardless of fiber position
                if rule["blend"].get("pure_first", False):
                    fiber_score += (max_positions + 1) * 1000
                else:
                    fiber_score += 500

            total_score += fiber_score

    return total_score

def find_standard_terms(feel_text):
    """Parse feel_terms text and return list of matched standard terms"""
    if not feel_text or not feel_text.strip():
        return []

    feel_input_lower = feel_text.lower()
    matched = []

    for standard_term, keywords in FEEL_DICTIONARY.items():
        for kw in sorted(keywords, key=len, reverse=True):
            if kw.lower() in feel_input_lower:
                if standard_term not in matched:
                    matched.append(standard_term)
                break

    return matched

def check_attribute_match(sample, attr_key, attr_rule):
    """Check if a single attribute matches the rule condition"""
    if attr_key in ("yarn", "blend", "weave", "finish"):
        values = attr_rule.get("values", [])
        exact_values = attr_rule.get("exact_values", [])
        if contains_any(sample[attr_key], values):
            return True
        for ev in exact_values:
            if sample[attr_key].lower().strip() == ev.lower().strip():
                return True
        return False
    elif attr_key == "count":
        if "min" in attr_rule and sample["count_avg"] < attr_rule["min"]:
            return False
        if "max" in attr_rule and sample["count_avg"] > attr_rule["max"]:
            return False
        return True
    elif attr_key == "gsm":
        if "min" in attr_rule and sample["gsm"] < attr_rule["min"]:
            return False
        if "max" in attr_rule and sample["gsm"] > attr_rule["max"]:
            return False
        return True
    return True

def check_combo_rule(sample, rule):
    """Check sample against combination-based rules.
    Returns (passes, score) where score indicates match quality.
    """
    best_score = 0
    for combo in rule.get("positive_combos", []):
        combo_score = combo["score"]
        match = True
        for attr in ("yarn", "blend", "weave"):
            if attr in combo:
                if not contains_any(sample[attr], combo[attr]):
                    match = False
                    break
        if match:
            best_score = max(best_score, combo_score)

    if best_score == 0:
        return False, 0

    # Add finish bonus
    finish_bonus = rule.get("finish_bonus", [])
    if finish_bonus and contains_any(sample["finish"], finish_bonus):
        best_score += 1

    return True, best_score

def check_sample_against_rule(sample, standard_term):
    """Check sample against rule with priority support.
    Returns (passes, score):
      - passes: False if any HIGH priority attribute fails or negative rules trigger
      - score: count of LOW priority attributes that match (for sorting)
    """
    rule = RULE_TABLE.get(standard_term)
    if not rule:
        return True, 0

    # Handle combination-based rules (e.g., Textured)
    if rule.get("combo_mode"):
        return check_combo_rule(sample, rule)

    # Check yarn_exclude (always hard reject)
    if rule.get("yarn_exclude"):
        if contains_any(sample["yarn"], rule["yarn_exclude"]):
            return False, 0

    # Check blend_exclude (always hard reject, unless blend_allow overrides)
    if rule.get("blend_exclude"):
        if contains_any(sample["blend"], rule["blend_exclude"]):
            if not contains_any(sample["blend"], rule.get("blend_allow", [])):
                return False, 0

    # Check negative cross-attribute rules (always hard reject)
    for neg in rule.get("negative_cross", []):
        all_match = True
        for attr, condition in neg.items():
            if attr.endswith("_override"):
                continue
            sample_val = sample.get(attr, "")
            if isinstance(condition, list):
                if not contains_any(sample_val, condition):
                    all_match = False
                    break
            elif isinstance(condition, str):
                if condition.lower() not in sample_val.lower():
                    all_match = False
                    break
        if all_match:
            # Check if an override cancels the rejection
            overridden = False
            for attr, condition in neg.items():
                if attr.endswith("_override"):
                    base_attr = attr.replace("_override", "")
                    sample_val = sample.get(base_attr, "")
                    if contains_any(sample_val, condition):
                        overridden = True
                        break
            if not overridden:
                return False, 0

    # Check each attribute with priority
    score = 0
    for attr_key in ("yarn", "count", "blend", "weave", "finish", "gsm"):
        if attr_key not in rule:
            continue
        attr_rule = rule[attr_key]
        if not isinstance(attr_rule, dict) or "priority" not in attr_rule:
            continue

        priority = attr_rule.get("priority")
        if priority is None:
            continue

        matches = check_attribute_match(sample, attr_key, attr_rule)

        if priority == "HIGH":
            if not matches:
                return False, 0
        elif priority == "LOW":
            if matches:
                score += 1

    return True, score

def filter_samples(product_type, gsm_min, gsm_max, blend, weave, yarn, feel_terms, finish="ALL"):
    """Apply all filters and return matching samples"""
    results = get_all_samples()

    # Step 1: Filter by direct parameters
    if product_type and product_type.upper() != "ALL":
        norm_filter = normalize_product(product_type)
        results = [s for s in results if normalize_product(s.get("product", "")) == norm_filter]

    if gsm_min:
        try:
            gsm_min_val = int(gsm_min)
            results = [s for s in results if int(s.get("gsm", 0)) >= gsm_min_val]
        except (ValueError, TypeError):
            pass

    if gsm_max:
        try:
            gsm_max_val = int(gsm_max)
            results = [s for s in results if int(s.get("gsm", 0)) <= gsm_max_val]
        except (ValueError, TypeError):
            pass

    if blend and blend.strip() and blend.upper() != "ALL":
        results = [s for s in results if blend.lower() in str(s.get("blend", "")).lower()]

    if weave and weave.upper() != "ALL":
        results = [s for s in results if weave.lower() in str(s.get("weave", "")).lower()]

    if yarn and yarn.upper() != "ALL":
        results = [s for s in results if yarn.lower() in str(s.get("yarn", "")).lower()]

    if finish and finish.strip() and finish.upper() != "ALL":
        results = [s for s in results if finish.lower() in str(s.get("finish", "")).lower()]

    # Step 2-4: Find standard terms and apply rules
    standard_terms = find_standard_terms(feel_terms)

    if feel_terms and feel_terms.strip() and not standard_terms:
        return [], standard_terms

    if standard_terms:
        for standard_term in standard_terms:
            filtered = []
            for s in results:
                passes, score = check_sample_against_rule(s, standard_term)
                if passes:
                    filtered.append(s)
            results = filtered

        # Calculate priority scores and sort best match first
        scored_results = []
        for s in results:
            s_copy = dict(s)
            s_copy["priority_score"] = calculate_priority_score(s, standard_terms)
            scored_results.append(s_copy)
        scored_results.sort(key=lambda s: -s["priority_score"])

        # Add recommendation rank
        for idx, s in enumerate(scored_results):
            s["rank"] = idx + 1
        results = scored_results

    # No filters applied (e.g. "All Products" with nothing else) -> show every sample
    return results, standard_terms

# ============================================================================
# FLASK ROUTES
# ============================================================================

IMAGES_DIR = os.path.join(BASE_DIR, "SAMPLE IMAGES")

def login_required(f):
    """Simple login check decorator."""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

@app.route("/sample-image/<int:sample_no>")
def sample_image(sample_no):
    filename = f"{sample_no}.jpeg"
    local_path = os.path.join(IMAGES_DIR, filename)
    if os.path.exists(local_path):
        return send_from_directory(IMAGES_DIR, filename)
    
    samples_2nd_dir = os.path.join(BASE_DIR, "public", "samples 2nd")
    if os.path.exists(os.path.join(samples_2nd_dir, filename)):
        return send_from_directory(samples_2nd_dir, filename)

    frontend_2nd_dir = os.path.join(BASE_DIR, "frontend", "public", "samples 2nd")
    if os.path.exists(os.path.join(frontend_2nd_dir, filename)):
        return send_from_directory(frontend_2nd_dir, filename)

    return redirect(f"{SUPABASE_URL}/storage/v1/object/public/Neha/{filename}")

@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_react(path):
    dist_dir = os.path.join(BASE_DIR, "frontend", "dist")
    file_path = os.path.join(dist_dir, path)
    if path and os.path.exists(file_path):
        return send_from_directory(dist_dir, path)
    index_path = os.path.join(dist_dir, "index.html")
    if os.path.exists(index_path):
        return send_from_directory(dist_dir, "index.html")
    if "user_id" in session:
        return redirect("/")
    return redirect(url_for("login"))


@app.route("/api/auth/me")
def api_auth_me():
    if "user_id" in session:
        buyer_info = session.get("buyer_info", {})
        return jsonify({
            "authenticated": True,
            "user": {
                "id": session["user_id"],
                "email": session.get("username", ""),
                **buyer_info
            }
        })
    return jsonify({"authenticated": False}), 401

@app.route("/api/auth/login", methods=["POST"])
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        is_json_req = request.is_json or request.headers.get("Accept") == "application/json" or request.path.startswith("/api/")
        if request.is_json:
            data = request.get_json() or {}
            email = data.get("email", "").strip()
            password = data.get("password", "")
        else:
            email = request.form.get("email", "").strip()
            password = request.form.get("password", "")
        try:
            result = sb_auth_signin(email, password)
            session["user_id"] = result["user"]["id"]
            session["username"] = result["user"]["email"]
            if is_json_req:
                return jsonify({"status": "ok", "user": {"id": result["user"]["id"], "email": result["user"]["email"]}})
            return redirect("/")
        except Exception as e:
            if is_json_req:
                return jsonify({"error": str(e) or "Invalid email or password"}), 400
            return render_template_string(LOGIN_TEMPLATE, error="Invalid email or password")
    if request.path.startswith("/api/"):
        return jsonify({"error": "Method not allowed"}), 405
    return render_template_string(LOGIN_TEMPLATE, error=None)

@app.route("/api/auth/register", methods=["POST"])
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        is_json_req = request.is_json or request.headers.get("Accept") == "application/json" or request.path.startswith("/api/")
        if request.is_json:
            data = request.get_json() or {}
            email = data.get("email", "").strip()
            password = data.get("password", "")
            confirm = data.get("confirm", password)
            buyer_name = data.get("buyerName", data.get("buyer_name", "")).strip()
            brand_name = data.get("brandName", data.get("brand_name", "")).strip()
            company = data.get("company", "").strip()
            country = data.get("country", "").strip()
            contact_person = data.get("contactPerson", data.get("contact_person", "")).strip()
            phone_number = data.get("phoneNumber", data.get("phone_number", "")).strip()
            buyer_id = data.get("buyerId", data.get("buyer_id", "")).strip()
        else:
            email = request.form.get("email", "").strip()
            password = request.form.get("password", "")
            confirm = request.form.get("confirm", "")
            buyer_name = request.form.get("buyer_name", "").strip()
            brand_name = request.form.get("brand_name", "").strip()
            company = request.form.get("company", "").strip()
            country = request.form.get("country", "").strip()
            contact_person = request.form.get("contact_person", "").strip()
            phone_number = request.form.get("phone_number", "").strip()
            buyer_id = request.form.get("buyer_id", "").strip()

        if not email or not password:
            err = "Email and password are required"
            if is_json_req:
                return jsonify({"error": err}), 400
            return render_template_string(REGISTER_TEMPLATE, error=err)
        if password != confirm:
            err = "Passwords do not match"
            if is_json_req:
                return jsonify({"error": err}), 400
            return render_template_string(REGISTER_TEMPLATE, error=err)
        if len(password) < 6:
            err = "Password must be at least 6 characters"
            if is_json_req:
                return jsonify({"error": err}), 400
            return render_template_string(REGISTER_TEMPLATE, error=err)
        try:
            result = sb_auth_signup(email, password)
            user = result.get("user")
            if user and user.get("id"):
                session["user_id"] = user["id"]
                session["username"] = user.get("email", email)
                session["buyer_info"] = {
                    "buyer_name": buyer_name,
                    "brand_name": brand_name,
                    "company": company,
                    "country": country,
                    "contact_person": contact_person,
                    "phone_number": phone_number,
                    "buyer_id": buyer_id,
                }
                # Automatically create a Wishlist Group for this buyer
                if buyer_name:
                    try:
                        existing = sb_select("wishlist_groups", filters={"user_id": f"eq.{user['id']}", "name": f"eq.{buyer_name}"})
                        if not existing:
                            sb_insert("wishlist_groups", {"user_id": user['id'], "name": buyer_name})
                    except Exception:
                        pass
                user_data = {
                    "id": user["id"],
                    "email": session["username"],
                    "buyer_name": buyer_name,
                    "brand_name": brand_name,
                    "company": company,
                    "country": country,
                    "contact_person": contact_person,
                    "phone_number": phone_number,
                    "buyer_id": buyer_id,
                }
                if is_json_req:
                    return jsonify({"status": "ok", "user": user_data})
                return redirect("/")
            if is_json_req:
                return jsonify({"error": "Registration failed"}), 400
            return render_template_string(REGISTER_TEMPLATE, error="Registration failed")
        except Exception as e:
            msg = str(e)
            err = "Email already registered" if "already" in msg.lower() else f"Error: {msg}"
            if is_json_req:
                return jsonify({"error": err}), 400
            return render_template_string(REGISTER_TEMPLATE, error=err)
    if request.path.startswith("/api/"):
        return jsonify({"error": "Method not allowed"}), 405
    return render_template_string(REGISTER_TEMPLATE, error=None)

@app.route("/api/auth/logout", methods=["POST", "GET"])
@app.route("/logout")
def logout():
    session.clear()
    if request.is_json or request.path.startswith("/api/"):
        return jsonify({"status": "logged_out"})
    return redirect(url_for("login"))

@app.route("/dashboard")
@login_required
def dashboard():
    return render_template_string(DASHBOARD_TEMPLATE, username=session.get("username", ""))

@app.route("/api/samples")
@login_required
def api_all_samples():
    samples = get_all_samples()
    return jsonify({"samples": samples, "total_count": len(samples)})

@app.route("/api/enquiries/create", methods=["POST"])
@login_required
def create_enquiry():
    data = request.get_json() or {}
    buyer_name = data.get("buyer_name", "").strip()
    if not buyer_name:
        return jsonify({"error": "Buyer Name is required"}), 400

    enquiries = session.get("enquiries", [])
    enquiry = {
        "enquiry_id": data.get("enquiry_id"),
        "buyer_name": buyer_name,
        "brand_name": data.get("brand_name"),
        "company": data.get("company"),
        "country": data.get("country"),
        "contact_person": data.get("contact_person"),
        "email": data.get("email"),
        "phone_number": data.get("phone_number"),
        "buyer_id": data.get("buyer_id"),
        "date_received": data.get("date_received"),
        "due_date": data.get("due_date"),
        "priority": data.get("priority", "High"),
        "requirement_type": data.get("requirement_type"),
        "season": data.get("season"),
        "quantity": data.get("quantity"),
        "status": data.get("status", "New"),
        "end_use": data.get("end_use"),
        "summary": data.get("summary"),
        "documents": data.get("documents", []),
        "created_at": "2025-07-22",
    }
    enquiries.append(enquiry)
    session["enquiries"] = enquiries

    # Also automatically create a wishlist group for this buyer if not existing
    try:
        user_id = session["user_id"]
        group_name = f"{buyer_name} ({data.get('season', 'Enquiry')})"
        existing = sb_select("wishlist_groups", filters={"user_id": f"eq.{user_id}", "name": f"eq.{group_name}"})
        if not existing:
            sb_insert("wishlist_groups", {"user_id": user_id, "name": group_name})
    except Exception:
        pass

    return jsonify({"status": "ok", "enquiry": enquiry})

@app.route("/api/enquiries")
@login_required
def get_enquiries():
    enquiries = session.get("enquiries", [])
    return jsonify({"enquiries": enquiries, "total_count": len(enquiries)})

@app.route("/api/enquiries/update", methods=["POST"])
@login_required
def update_enquiry():
    data = request.get_json() or {}
    enquiry_id = data.get("enquiry_id")
    if not enquiry_id:
        return jsonify({"error": "enquiry_id is required"}), 400

    enquiries = session.get("enquiries", [])
    updated = False
    for i, enq in enumerate(enquiries):
        if enq.get("enquiry_id") == enquiry_id:
            enquiries[i].update(data)
            updated = True
            break

    if not updated:
        enquiries.append(data)

    session["enquiries"] = enquiries
    return jsonify({"status": "ok", "enquiry_id": enquiry_id})

@app.route("/search", methods=["POST"])
@app.route("/api/search", methods=["POST"])
@login_required
def search():
    if request.is_json:
        data = request.get_json() or {}
        product_type = data.get("product_type", "ALL")
        gsm_min = str(data.get("gsm_min", "")).strip()
        gsm_max = str(data.get("gsm_max", "")).strip()
        blend = str(data.get("blend", "")).strip()
        weave = data.get("weave", "ALL")
        yarn = data.get("yarn", "ALL")
        finish = str(data.get("finish_type", data.get("finish", "ALL"))).strip()
        feel_terms = str(data.get("feel_terms", "")).strip()
    else:
        product_type = request.form.get("product_type", "ALL")
        gsm_min = request.form.get("gsm_min", "").strip()
        gsm_max = request.form.get("gsm_max", "").strip()
        blend = request.form.get("blend", "").strip()
        weave = request.form.get("weave", "ALL")
        yarn = request.form.get("yarn", "ALL")
        finish = request.form.get("finish_type", request.form.get("finish", "ALL")).strip()
        feel_terms = request.form.get("feel_terms", "").strip()

    results, standard_terms = filter_samples(product_type, gsm_min, gsm_max, blend, weave, yarn, feel_terms, finish)

    return jsonify({
        "results": results,
        "standard_terms": standard_terms,
        "total_count": len(results),
    })


@app.route("/api/wishlist")
@login_required
def get_wishlist():
    user_id = session["user_id"]
    groups = sb_select("wishlist_groups", filters={"user_id": f"eq.{user_id}"}, order="name")
    result = []
    all_sample_nos = []
    for g in groups:
        items = sb_select("wishlists", columns="sample_no", filters={"user_id": f"eq.{user_id}", "group_id": f"eq.{g['id']}"})
        sample_nos = [item["sample_no"] for item in items]
        all_sample_nos.extend(sample_nos)
        if sample_nos:
            in_list = ",".join(str(n) for n in sample_nos)
            samples = sb_select("samples", filters={"sample_no": f"in.({in_list})"})
        else:
            samples = []
        result.append({"group_id": g["id"], "group_name": g["name"], "samples": samples, "count": len(samples)})
    return jsonify({"groups": result, "all_sample_nos": all_sample_nos, "total_count": len(all_sample_nos)})

@app.route("/api/wishlist/groups", methods=["GET"])
@login_required
def get_wishlist_groups():
    user_id = session["user_id"]
    groups = sb_select("wishlist_groups", columns="id,name", filters={"user_id": f"eq.{user_id}"}, order="name")
    return jsonify({"groups": groups})

@app.route("/api/wishlist/groups/create", methods=["POST"])
@login_required
def create_wishlist_group():
    data = request.get_json()
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Group name is required"}), 400
    user_id = session["user_id"]
    existing = sb_select("wishlist_groups", columns="id", filters={"user_id": f"eq.{user_id}", "name": f"eq.{name}"})
    if existing:
        return jsonify({"error": "Group already exists"}), 400
    result = sb_insert("wishlist_groups", {"user_id": user_id, "name": name})
    group = result[0]
    return jsonify({"status": "created", "group_id": group["id"], "name": group["name"]})

@app.route("/api/wishlist/groups/delete", methods=["POST"])
@login_required
def delete_wishlist_group():
    data = request.get_json()
    group_id = data.get("group_id")
    user_id = session["user_id"]
    group = sb_select("wishlist_groups", columns="id", filters={"id": f"eq.{group_id}", "user_id": f"eq.{user_id}"})
    if not group:
        return jsonify({"error": "Group not found"}), 404
    sb_delete("wishlists", {"user_id": f"eq.{user_id}", "group_id": f"eq.{group_id}"})
    sb_delete("wishlist_groups", {"id": f"eq.{group_id}", "user_id": f"eq.{user_id}"})
    return jsonify({"status": "deleted"})

@app.route("/api/wishlist/add", methods=["POST"])
@login_required
def add_to_wishlist():
    data = request.get_json()
    sample_no = data.get("sample_no")
    group_id = data.get("group_id")
    if not sample_no or not group_id:
        return jsonify({"error": "sample_no and group_id required"}), 400
    user_id = session["user_id"]
    group = sb_select("wishlist_groups", columns="id", filters={"id": f"eq.{group_id}", "user_id": f"eq.{user_id}"})
    if not group:
        return jsonify({"error": "Group not found"}), 404
    existing = sb_select("wishlists", columns="id", filters={"user_id": f"eq.{user_id}", "group_id": f"eq.{group_id}", "sample_no": f"eq.{sample_no}"})
    if not existing:
        sb_insert("wishlists", {"user_id": user_id, "group_id": group_id, "sample_no": sample_no})
    return jsonify({"status": "added", "sample_no": sample_no})

@app.route("/api/wishlist/remove", methods=["POST"])
@login_required
def remove_from_wishlist():
    data = request.get_json()
    sample_no = data.get("sample_no")
    group_id = data.get("group_id")
    if not sample_no or not group_id:
        return jsonify({"error": "sample_no and group_id required"}), 400
    user_id = session["user_id"]
    sb_delete("wishlists", {"user_id": f"eq.{user_id}", "group_id": f"eq.{group_id}", "sample_no": f"eq.{sample_no}"})
    return jsonify({"status": "removed", "sample_no": sample_no})

# ============================================================================
# EXCEL UPLOAD — bulk-add samples from a transposed sheet with embedded images
# ============================================================================
STORAGE_BUCKET = "Neha"

# Maps a row-label found in column A to our record field name
_LABEL_TO_FIELD = {
    "ARTICLE": "article",
    "SAMPLE NO": "sample_no", "SAMPLE NUMBER": "sample_no", "SAMPLE": "sample_no",
    "PRODUCT": "product",
    "YARN": "yarn",
    "COUNT": "count",
    "CONSTRUCTION": "construction",
    "BLEND": "blend",
    "WEAVE": "weave",
    "FINISH": "finish",
    "GSM": "gsm",
}

def _clean_text(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()

def _clean_dim(v):
    """Normalise an 'A*B' dimension string (count / construction): fix OCR 'O'->'0', unify separator."""
    s = _clean_text(v).upper().replace(" ", "")
    return s.replace("O", "0").replace("X", "*").replace("×", "*")

def _split_pair(s):
    nums = []
    for p in re.split(r"[*]", s):
        digits = re.sub(r"[^0-9]", "", p)
        if digits:
            nums.append(int(digits))
    return nums

def parse_samples_workbook(file_bytes):
    """Parse a transposed sheet (attributes as rows, one sample per column).
    Returns (records, images, warnings); images maps sample_no -> JPEG bytes."""
    import openpyxl
    from PIL import Image as PILImage

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # 1) Locate the row that holds each field by reading the label column (col A)
    label_row = {}
    for r in range(1, ws.max_row + 1):
        raw = ws.cell(row=r, column=1).value
        if raw is None:
            continue
        key = _clean_text(raw).upper().rstrip(".")
        field = _LABEL_TO_FIELD.get(key)
        if field:
            label_row[field] = r

    if "sample_no" not in label_row:
        raise ValueError("Could not find a 'SAMPLE NO.' row in the first column of the sheet.")

    warnings = []

    # 2) Sample columns = columns where the SAMPLE NO. row holds a number
    sno_row = label_row["sample_no"]
    col_to_sno = {}  # 0-indexed column -> sample_no (0-index aligns with image anchors)
    for c in range(2, ws.max_column + 1):
        val = ws.cell(row=sno_row, column=c).value
        if val is None or _clean_text(val) == "":
            continue
        try:
            col_to_sno[c - 1] = int(float(str(val).strip()))
        except ValueError:
            continue

    # 3) Build one record per sample column
    records = []
    for col0, sno in sorted(col_to_sno.items()):
        c = col0 + 1

        def cell(field):
            r = label_row.get(field)
            return ws.cell(row=r, column=c).value if r else None

        count_s = _clean_dim(cell("count"))
        cons_s = _clean_dim(cell("construction"))
        cnums = _split_pair(count_s)
        pnums = _split_pair(cons_s)

        try:
            gsm_raw = cell("gsm")
            gsm = int(float(str(gsm_raw).strip())) if gsm_raw not in (None, "") else 0
        except ValueError:
            gsm = 0
            warnings.append(f"Sample {sno}: GSM '{gsm_raw}' not numeric -> 0")

        records.append({
            "sample_no": sno,
            "article": _clean_text(cell("article")),
            "product": _clean_text(cell("product")),
            "yarn": _clean_text(cell("yarn")),
            "count": count_s,
            "count_avg": sum(cnums) // len(cnums) if cnums else 0,
            "construction": cons_s,
            "construction_total": sum(pnums) if pnums else 0,
            "blend": _clean_text(cell("blend")),
            "weave": _clean_text(cell("weave")),
            "finish": _clean_text(cell("finish")),
            "gsm": gsm,
        })

    # 4) Extract embedded images, match each to a sample column by its anchor
    images = {}
    for im in getattr(ws, "_images", []):
        try:
            col0 = im.anchor._from.col
        except Exception:
            continue
        sno = col_to_sno.get(col0)
        if sno is None and col_to_sno:
            sno = min(col_to_sno.items(), key=lambda kv: abs(kv[0] - col0))[1]
        if sno is None:
            continue
        try:
            pil = PILImage.open(io.BytesIO(im._data()))
            if pil.mode != "RGB":
                pil = pil.convert("RGB")
            buf = io.BytesIO()
            pil.save(buf, format="JPEG", quality=88)
            images[sno] = buf.getvalue()
        except Exception as e:
            warnings.append(f"Sample {sno}: embedded image unreadable ({e})")

    return records, images, warnings

def sb_upsert_samples(records):
    """Insert or update sample rows keyed on the unique sample_no column."""
    headers = _sb_headers({"Prefer": "resolution=merge-duplicates,return=minimal"})
    r = httpx.post(f"{SUPABASE_URL}/rest/v1/samples?on_conflict=sample_no",
                   json=records, headers=headers, timeout=60.0)
    r.raise_for_status()

def sb_upload_image(sample_no, jpeg_bytes):
    headers = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}",
               "Content-Type": "image/jpeg", "x-upsert": "true"}
    r = httpx.post(f"{SUPABASE_URL}/storage/v1/object/{STORAGE_BUCKET}/{sample_no}.jpeg",
                   content=jpeg_bytes, headers=headers, timeout=120.0)
    return r.status_code in (200, 201)

@app.route("/api/upload-samples", methods=["POST"])
@login_required
def upload_samples():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "No file uploaded"}), 400
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Please upload an Excel .xlsx file"}), 400
    try:
        records, images, warnings = parse_samples_workbook(f.read())
    except Exception as e:
        return jsonify({"error": f"Could not read the sheet: {e}"}), 400

    if not records:
        return jsonify({"error": "No samples found in the sheet."}), 400

    try:
        sb_upsert_samples(records)
    except Exception as e:
        return jsonify({"error": f"Saving to database failed: {e}"}), 500

    img_ok = img_fail = 0
    for sno, data in images.items():
        if sb_upload_image(sno, data):
            img_ok += 1
        else:
            img_fail += 1
            warnings.append(f"Sample {sno}: image upload failed")

    sample_nos = sorted(r["sample_no"] for r in records)
    return jsonify({
        "status": "ok",
        "samples_added": len(records),
        "sample_range": [sample_nos[0], sample_nos[-1]],
        "images_uploaded": img_ok,
        "images_failed": img_fail,
        "no_image": [s for s in sample_nos if s not in images],
        "warnings": warnings,
    })

_REQUIRED_FIELDS = ["article", "product", "yarn", "count", "blend", "weave", "finish"]

@app.route("/api/upload-samples/preview", methods=["POST"])
@login_required
def preview_samples():
    """Parse the sheet and return a reviewable preview WITHOUT writing anything."""
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "No file uploaded"}), 400
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Please upload an Excel .xlsx file"}), 400
    try:
        records, images, warnings = parse_samples_workbook(f.read())
    except Exception as e:
        return jsonify({"error": f"Could not read the sheet: {e}"}), 400
    if not records:
        return jsonify({"error": "No samples found in the sheet."}), 400

    try:
        existing = {r["sample_no"] for r in sb_select("samples", columns="sample_no")}
    except Exception:
        existing = set()

    import base64
    from PIL import Image as PILImage
    thumbs = {}
    for sno, data in images.items():
        try:
            pil = PILImage.open(io.BytesIO(data))
            pil.thumbnail((80, 80))
            b = io.BytesIO()
            pil.save(b, "JPEG", quality=72)
            thumbs[sno] = "data:image/jpeg;base64," + base64.b64encode(b.getvalue()).decode()
        except Exception:
            pass

    # detect duplicate sample numbers within the sheet itself
    seen, dupes = set(), set()
    for r in records:
        if r["sample_no"] in seen:
            dupes.add(r["sample_no"])
        seen.add(r["sample_no"])

    out = []
    for r in records:
        issues = [fld for fld in _REQUIRED_FIELDS if not r.get(fld)]
        if r["gsm"] == 0:
            issues.append("gsm")
        if r["count_avg"] == 0:
            issues.append("count")
        if r["sample_no"] in dupes:
            issues.append("duplicate")
        rec = dict(r)
        rec["status"] = "update" if r["sample_no"] in existing else "new"
        rec["has_image"] = r["sample_no"] in images
        rec["thumb"] = thumbs.get(r["sample_no"])
        rec["issues"] = issues
        out.append(rec)

    return jsonify({
        "status": "ok",
        "records": out,
        "summary": {
            "total": len(out),
            "new": sum(1 for r in out if r["status"] == "new"),
            "updates": sum(1 for r in out if r["status"] == "update"),
            "with_image": len(images),
            "without_image": len(out) - len(images),
            "with_issues": sum(1 for r in out if r["issues"]),
        },
        "warnings": warnings,
    })

# ============================================================================
# HTML TEMPLATES
# ============================================================================

_AUTH_STYLES = """
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@600;700&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
    * { margin: 0; padding: 0; box-sizing: border-box; }
    body {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
        background: linear-gradient(160deg, #0D1B2A 0%, #1B2A4A 45%, #2C3E6B 100%);
        min-height: 100vh;
        display: flex;
        align-items: center;
        justify-content: center;
        position: relative;
        overflow: hidden;
    }
    body::before {
        content: '';
        position: absolute; inset: 0;
        background-image:
            repeating-linear-gradient(45deg, transparent, transparent 30px, rgba(196,149,42,0.035) 30px, rgba(196,149,42,0.035) 31px),
            repeating-linear-gradient(-45deg, transparent, transparent 30px, rgba(196,149,42,0.035) 30px, rgba(196,149,42,0.035) 31px);
        pointer-events: none;
    }
    body::after {
        content: '';
        position: absolute;
        width: 500px; height: 500px;
        background: radial-gradient(circle, rgba(196,149,42,0.12) 0%, transparent 70%);
        top: -100px; right: -100px;
        pointer-events: none;
    }
    .auth-container {
        width: 100%; max-width: 440px; padding: 1.5rem; position: relative; z-index: 1;
    }
    .auth-logo {
        text-align: center; margin-bottom: 2.2rem; color: white;
    }
    .auth-logo-icon {
        width: 66px; height: 66px;
        background: linear-gradient(135deg, #C4952A, #E8C565);
        border-radius: 18px;
        display: inline-flex; align-items: center; justify-content: center;
        font-size: 2rem; margin-bottom: 1rem;
        box-shadow: 0 8px 28px rgba(196,149,42,0.45);
    }
    .auth-logo h1 {
        font-family: 'Playfair Display', serif;
        font-size: 2rem; letter-spacing: 0.5px; margin-bottom: 0.4rem;
        background: linear-gradient(135deg, #E8C565 0%, #ffffff 60%);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    .auth-logo p { opacity: 0.6; font-size: 0.88rem; color: #c8d4f0; letter-spacing: 0.4px; }
    .auth-card {
        background: #FFFDF8;
        border-radius: 20px;
        padding: 2.6rem 2.4rem;
        box-shadow: 0 28px 70px rgba(0,0,0,0.45), 0 0 0 1px rgba(196,149,42,0.18);
        border-top: 4px solid #C4952A;
        position: relative;
        overflow: hidden;
    }
    .auth-card::before {
        content: '';
        position: absolute; top: 0; right: 0;
        width: 120px; height: 120px;
        background: radial-gradient(circle at top right, rgba(196,149,42,0.08) 0%, transparent 65%);
    }
    .auth-card h2 {
        font-family: 'Playfair Display', serif;
        color: #1B2A4A; margin-bottom: 1.6rem; font-size: 1.55rem; text-align: center;
        letter-spacing: 0.3px;
    }
    .auth-field { margin-bottom: 1.25rem; }
    .auth-field label {
        display: block; font-weight: 600; color: #1B2A4A;
        font-size: 0.75rem; margin-bottom: 0.45rem;
        text-transform: uppercase; letter-spacing: 0.8px;
    }
    .auth-field input {
        width: 100%;
        padding: 0.88rem 1.1rem;
        border: 1.5px solid #DDD5C5;
        border-radius: 10px;
        font-size: 0.97rem;
        font-family: 'Inter', sans-serif;
        transition: all 0.25s;
        background: white;
        color: #2C2C2C;
    }
    .auth-field input::placeholder { color: #B8B0A0; }
    .auth-field input:focus {
        outline: none; border-color: #C4952A;
        box-shadow: 0 0 0 3px rgba(196,149,42,0.15);
        background: white;
    }
    .auth-btn {
        width: 100%;
        padding: 0.92rem;
        background: linear-gradient(135deg, #8B1535 0%, #B01C43 60%, #C4952A 100%);
        color: white;
        border: none;
        border-radius: 10px;
        font-size: 1rem;
        font-weight: 700;
        font-family: 'Inter', sans-serif;
        cursor: pointer;
        transition: all 0.3s;
        margin-top: 0.8rem;
        letter-spacing: 0.5px;
        text-transform: uppercase;
    }
    .auth-btn:hover {
        transform: translateY(-2px);
        box-shadow: 0 10px 28px rgba(139,21,53,0.45);
        background: linear-gradient(135deg, #A01A3E 0%, #C4952A 100%);
    }
    .auth-btn:active { transform: translateY(0); }
    .auth-error {
        background: #fff0f0; border: 1px solid #ffcdd2;
        color: #c62828; padding: 0.7rem 1rem; border-radius: 8px;
        margin-bottom: 1rem; font-size: 0.9rem; text-align: center;
    }
    .auth-link {
        text-align: center; margin-top: 1.6rem; font-size: 0.9rem; color: #7A6E5E;
    }
    .auth-link a { color: #C4952A; text-decoration: none; font-weight: 600; }
    .auth-link a:hover { text-decoration: underline; }
    .auth-divider {
        display: flex; align-items: center; gap: 1rem; margin: 1rem 0 0.5rem;
        color: #B8B0A0; font-size: 0.78rem;
    }
    .auth-divider::before, .auth-divider::after {
        content: ''; flex: 1; height: 1px; background: #E8E0D0;
    }
</style>
"""

LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Sign In — Fabric Sample Tool</title>
    """ + _AUTH_STYLES + """
</head>
<body>
    <div class="auth-container">
        <div class="auth-logo">
            <div class="auth-logo-icon">&#129525;</div>
            <h1>FabricSample</h1>
            <p>Smart Search by Buyer Requirements</p>
        </div>
        <div class="auth-card">
            <h2>Welcome Back</h2>
            {% if error %}
            <div class="auth-error">{{ error }}</div>
            {% endif %}
            <form method="POST">
                <div class="auth-field">
                    <label for="email">Email Address</label>
                    <input type="email" id="email" name="email" placeholder="you@example.com" required autofocus>
                </div>
                <div class="auth-field">
                    <label for="password">Password</label>
                    <input type="password" id="password" name="password" placeholder="••••••••" required>
                </div>
                <button type="submit" class="auth-btn">Sign In</button>
            </form>
            <div class="auth-divider">or</div>
            <div class="auth-link">
                Don't have an account? <a href="/register">Create one &rarr;</a>
            </div>
        </div>
    </div>
</body>
</html>
"""

REGISTER_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Register — Fabric Sample Tool</title>
    """ + _AUTH_STYLES + """
</head>
<body>
    <div class="auth-container">
        <div class="auth-logo">
            <div class="auth-logo-icon">&#129525;</div>
            <h1>FabricSample</h1>
            <p>Create your account to get started</p>
        </div>
        <div class="auth-card">
            <h2>Create Account</h2>
            {% if error %}
            <div class="auth-error">{{ error }}</div>
            {% endif %}
            <form method="POST">
                <div class="auth-field">
                    <label for="email">Email Address</label>
                    <input type="email" id="email" name="email" placeholder="you@example.com" required autofocus>
                </div>
                <div class="auth-field">
                    <label for="password">Password</label>
                    <input type="password" id="password" name="password" placeholder="Create a strong password" required>
                </div>
                <div class="auth-field">
                    <label for="confirm">Confirm Password</label>
                    <input type="password" id="confirm" name="confirm" placeholder="Confirm your password" required>
                </div>
                <button type="submit" class="auth-btn">Create Account</button>
            </form>
            <div class="auth-divider">or</div>
            <div class="auth-link">
                Already have an account? <a href="/login">Sign in &rarr;</a>
            </div>
        </div>
    </div>
</body>
</html>
"""

DASHBOARD_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Dashboard — Fabric Sample Tool</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@600;700&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        :root {
            --navy:    #1B2A4A;
            --navy2:   #243659;
            --burgundy:#8B1535;
            --gold:    #C4952A;
            --gold-lt: #EDD78A;
            --cream:   #F8F3E8;
            --warm-wh: #FFFDF8;
            --charcoal:#2D2D2D;
            --muted:   #7A6E5E;
            --border:  #E4DDD0;
            --teal:    #0097a7;
        }
        * { margin: 0; padding: 0; box-sizing: border-box; }

        body {
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
            background-color: var(--cream);
            color: var(--charcoal);
            line-height: 1.6;
        }

        /* ========== HEADER ========== */
        .header {
            background: linear-gradient(135deg, var(--navy) 0%, var(--navy2) 100%);
            color: white;
            padding: 0 2rem;
            height: 64px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            box-shadow: 0 3px 16px rgba(0,0,0,0.25);
            position: sticky;
            top: 0;
            z-index: 200;
            border-bottom: 2px solid var(--gold);
        }
        .header-brand {
            display: flex; align-items: center; gap: 0.9rem;
        }
        .header-brand-icon {
            width: 38px; height: 38px;
            background: linear-gradient(135deg, var(--gold), var(--gold-lt));
            border-radius: 10px;
            display: flex; align-items: center; justify-content: center;
            font-size: 1.3rem;
            box-shadow: 0 3px 10px rgba(196,149,42,0.4);
            flex-shrink: 0;
        }
        .header-brand-text h1 {
            font-family: 'Playfair Display', serif;
            font-size: 1.25rem; letter-spacing: 0.3px; color: white; line-height: 1.1;
        }
        .header-brand-text p { font-size: 0.72rem; opacity: 0.55; color: var(--gold-lt); letter-spacing: 0.3px; }
        .header-right { display: flex; align-items: center; gap: 1rem; }
        .header-user {
            display: flex; align-items: center; gap: 0.65rem;
            background: rgba(255,255,255,0.08);
            padding: 0.45rem 1rem;
            border-radius: 10px;
            border: 1px solid rgba(196,149,42,0.25);
        }
        .header-avatar {
            width: 32px; height: 32px;
            background: linear-gradient(135deg, var(--gold), var(--gold-lt));
            border-radius: 50%;
            display: flex; align-items: center; justify-content: center;
            font-weight: 700; font-size: 0.85rem; color: var(--navy);
        }
        .header-username { font-weight: 600; font-size: 0.9rem; }
        .btn-logout {
            padding: 0.45rem 1.1rem;
            background: rgba(139,21,53,0.3);
            color: white;
            border: 1px solid rgba(139,21,53,0.5);
            border-radius: 8px;
            font-size: 0.82rem; font-weight: 600;
            cursor: pointer; text-decoration: none;
            transition: all 0.25s; font-family: 'Inter', sans-serif;
            letter-spacing: 0.3px;
        }
        .btn-logout:hover { background: rgba(139,21,53,0.6); border-color: var(--burgundy); }

        /* ========== SIDEBAR ========== */
        .app-layout { display: flex; min-height: calc(100vh - 64px); }

        .sidebar {
            width: 230px;
            background: var(--navy);
            padding: 1.8rem 0;
            flex-shrink: 0;
            position: sticky;
            top: 64px;
            height: calc(100vh - 64px);
            overflow-y: auto;
            box-shadow: 3px 0 12px rgba(0,0,0,0.18);
        }
        .sidebar-label {
            padding: 0 1.5rem;
            font-size: 0.65rem; font-weight: 700;
            text-transform: uppercase; letter-spacing: 1.5px;
            color: rgba(196,149,42,0.6); margin-bottom: 0.6rem;
        }
        .nav-item {
            display: flex; align-items: center; gap: 0.85rem;
            padding: 0.85rem 1.5rem;
            color: rgba(255,255,255,0.6); font-weight: 500; font-size: 0.92rem;
            cursor: pointer; transition: all 0.2s;
            border-left: 3px solid transparent;
            position: relative;
        }
        .nav-item:hover { background: rgba(255,255,255,0.06); color: var(--gold-lt); }
        .nav-item.active {
            background: rgba(196,149,42,0.12);
            color: var(--gold-lt);
            border-left-color: var(--gold);
            font-weight: 600;
        }
        .nav-icon { font-size: 1.15rem; width: 22px; text-align: center; }
        .nav-badge {
            margin-left: auto;
            background: var(--burgundy); color: white;
            font-size: 0.68rem; padding: 0.12rem 0.5rem;
            border-radius: 10px; font-weight: 700;
        }
        .sidebar-divider {
            height: 1px; background: rgba(255,255,255,0.07);
            margin: 1rem 1.5rem;
        }

        @media (max-width: 768px) {
            .sidebar { width: 56px; }
            .nav-item span:not(.nav-icon) { display: none; }
            .nav-badge { display: none; }
            .sidebar-label { display: none; }
            .nav-item { justify-content: center; padding: 0.85rem; }
        }

        /* ========== MAIN CONTENT ========== */
        .main-content { flex: 1; padding: 2rem; overflow-y: auto; min-width: 0; }
        .page-section { display: none; }
        .page-section.active { display: block; }
        .container { max-width: 1120px; margin: 0 auto; }

        /* Section title bar */
        .section-title {
            display: flex; align-items: center; gap: 0.8rem;
            margin-bottom: 1.5rem;
        }
        .section-title h2 {
            font-family: 'Playfair Display', serif;
            color: var(--navy); font-size: 1.45rem;
        }
        .section-title-line {
            flex: 1; height: 2px;
            background: linear-gradient(90deg, var(--gold), transparent);
        }

        /* ========== CARD ========== */
        .card {
            background: var(--warm-wh);
            border-radius: 14px;
            padding: 2rem;
            box-shadow: 0 2px 14px rgba(0,0,0,0.07);
            margin-bottom: 2rem;
            border: 1px solid var(--border);
            border-top: 3px solid var(--gold);
        }
        .card-title {
            font-family: 'Playfair Display', serif;
            color: var(--navy); font-size: 1.2rem;
            margin-bottom: 1.5rem;
            display: flex; align-items: center; gap: 0.6rem;
        }
        .card-title::after {
            content: ''; display: inline-block;
            width: 32px; height: 2px;
            background: var(--gold); border-radius: 2px;
        }

        /* ========== FORM ========== */
        .form-grid {
            display: grid;
            grid-template-columns: 1fr 1fr 1fr;
            gap: 1.2rem; margin-bottom: 1.2rem;
        }
        @media (max-width: 900px) { .form-grid { grid-template-columns: 1fr 1fr; } }
        @media (max-width: 600px) { .form-grid { grid-template-columns: 1fr; } }

        .form-group { display: flex; flex-direction: column; }
        .form-group.full-width { grid-column: 1 / -1; }

        label {
            font-weight: 600; margin-bottom: 0.4rem; color: var(--navy);
            font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.7px;
        }

        input[type="text"], input[type="number"], select {
            padding: 0.72rem 0.95rem;
            border: 1.5px solid var(--border); border-radius: 9px;
            font-size: 0.93rem; font-family: 'Inter', sans-serif;
            transition: all 0.25s;
            background: white;
            color: var(--charcoal);
        }
        input[type="text"]::placeholder, input[type="number"]::placeholder { color: #C0B8A8; }
        input:focus, select:focus {
            outline: none; border-color: var(--gold);
            box-shadow: 0 0 0 3px rgba(196,149,42,0.14); background: white;
        }

        .button-group { display: flex; gap: 1rem; justify-content: flex-end; margin-top: 1.6rem; }

        button {
            padding: 0.72rem 1.6rem; font-size: 0.92rem; font-weight: 600;
            border: none; border-radius: 9px; cursor: pointer;
            transition: all 0.25s; font-family: 'Inter', sans-serif;
        }
        .btn-search {
            background: linear-gradient(135deg, var(--burgundy), #A01A3E);
            color: white; letter-spacing: 0.4px;
        }
        .btn-search:hover { transform: translateY(-2px); box-shadow: 0 8px 20px rgba(139,21,53,0.38); }
        .btn-reset { background: #EDE8DF; color: var(--muted); border: 1px solid var(--border); }
        .btn-reset:hover { background: #E0D8CE; color: var(--navy); }

        /* ========== RESULTS ========== */
        .results-section { display: none; }
        .results-section.show { display: block; }
        .results-info {
            background: linear-gradient(135deg, #FFF8E8, #FFF3D0);
            border-left: 4px solid var(--gold);
            padding: 1rem 1.3rem; margin-bottom: 1.5rem; border-radius: 10px;
            border: 1px solid rgba(196,149,42,0.2);
        }
        .results-info h3 { color: var(--navy); margin-bottom: 0.3rem; font-family: 'Playfair Display', serif; }
        .results-info p { color: var(--muted); font-size: 0.88rem; }
        .tags { display: flex; flex-wrap: wrap; gap: 0.45rem; margin-top: 0.5rem; }
        .tag {
            background: linear-gradient(135deg, var(--navy), var(--navy2));
            color: var(--gold-lt);
            padding: 0.22rem 0.8rem; border-radius: 20px;
            font-size: 0.78rem; font-weight: 600; letter-spacing: 0.3px;
        }

        /* ========== SAMPLE CARDS ========== */
        .samples-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(265px, 1fr));
            gap: 1.5rem;
        }
        .sample-card {
            background: var(--warm-wh);
            border-radius: 14px; overflow: hidden;
            box-shadow: 0 2px 12px rgba(0,0,0,0.08);
            transition: transform 0.25s, box-shadow 0.25s;
            position: relative; cursor: pointer;
            border: 1px solid var(--border);
        }
        .sample-card::before {
            content: '';
            position: absolute; top: 0; left: 0; right: 0; height: 3px;
            background: linear-gradient(90deg, var(--gold), var(--burgundy));
            z-index: 1;
        }
        .sample-card:hover {
            transform: translateY(-5px);
            box-shadow: 0 12px 32px rgba(0,0,0,0.14);
        }
        .sample-card-img {
            width: 100%; height: 195px;
            object-fit: cover; display: block;
            background: linear-gradient(135deg, #EDE8DF, #DDD5C5);
        }
        .sample-card-body { padding: 1rem 1.1rem 1.15rem; }
        .sample-card-header {
            display: flex; justify-content: space-between; align-items: center;
            margin-bottom: 0.65rem; padding-bottom: 0.55rem;
            border-bottom: 1px solid var(--border);
        }
        .sample-card-no {
            font-weight: 700; font-size: 1rem;
            color: var(--navy); font-family: 'Playfair Display', serif;
        }
        .sample-card-article {
            font-size: 0.78rem; color: var(--gold);
            font-weight: 600; letter-spacing: 0.3px;
        }
        .sample-card-props {
            display: grid; grid-template-columns: 1fr 1fr;
            gap: 0.35rem 0.8rem;
        }
        .sample-prop { font-size: 0.8rem; }
        .sample-prop-label {
            color: #B0A898; font-weight: 700;
            text-transform: uppercase; font-size: 0.63rem; letter-spacing: 0.7px;
        }
        .sample-prop-value { color: var(--charcoal); font-weight: 500; font-size: 0.82rem; }

        /* ========== RANK BADGE ========== */
        .rank-badge {
            position: absolute; top: 12px; left: 10px;
            padding: 0.22rem 0.65rem; border-radius: 8px;
            font-weight: 700; font-size: 0.75rem; z-index: 3;
            letter-spacing: 0.3px;
        }
        .rank-top { background: linear-gradient(135deg, #2e7d32, #388e3c); color: white; }
        .rank-mid { background: linear-gradient(135deg, #e65100, #f57c00); color: white; }
        .rank-normal { background: rgba(255,255,255,0.92); color: #555; }
        .rank-label {
            position: absolute; top: 40px; left: 10px;
            font-size: 0.6rem; color: white;
            background: rgba(0,0,0,0.5);
            padding: 0.1rem 0.45rem; border-radius: 4px; z-index: 3;
            letter-spacing: 0.3px;
        }

        /* ========== PAGE HEADER ========== */
        .page-header {
            display: flex; justify-content: space-between; align-items: center;
            margin-bottom: 1.5rem; flex-wrap: wrap; gap: 1rem;
        }
        .page-header h2 {
            font-family: 'Playfair Display', serif;
            color: var(--navy); font-size: 1.4rem;
        }
        .count-badge {
            background: linear-gradient(135deg, var(--navy), var(--navy2));
            color: var(--gold-lt);
            padding: 0.35rem 1.1rem; border-radius: 20px;
            font-weight: 600; font-size: 0.82rem;
        }

        /* ========== EMPTY STATE ========== */
        .empty-state { text-align: center; padding: 4rem 2rem; color: #B0A898; }
        .empty-state .empty-icon { font-size: 3.2rem; margin-bottom: 1rem; }
        .empty-state p { font-size: 1.05rem; font-weight: 600; color: var(--muted); }
        .empty-state .sub { font-size: 0.88rem; margin-top: 0.5rem; font-weight: 400; }

        /* ========== LOADING ========== */
        .loading { display: none; text-align: center; padding: 3rem; color: var(--gold); }
        .loading.show { display: block; }
        .spinner {
            border: 3px solid rgba(196,149,42,0.2); border-top: 3px solid var(--gold);
            border-radius: 50%; width: 38px; height: 38px;
            animation: spin 0.8s linear infinite; margin: 0 auto 1rem;
        }
        .loading p { color: var(--muted); font-size: 0.9rem; }
        @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }

        .no-results { text-align: center; padding: 3rem 1rem; color: #B0A898; }
        .no-results p { font-size: 1.05rem; font-weight: 500; }

        .error-message {
            background: #fff0f0; border-left: 4px solid #f44336;
            padding: 1rem; margin-bottom: 1.5rem; border-radius: 8px;
            color: #c62828; display: none;
        }
        .error-message.show { display: block; }

        /* ========== WISHLIST GROUPS ========== */
        .group-bar {
            display: flex; gap: 0.8rem; align-items: center;
            margin-bottom: 1.5rem; flex-wrap: wrap;
        }
        .group-chip {
            padding: 0.48rem 1.2rem;
            border-radius: 20px; font-size: 0.88rem; font-weight: 600;
            cursor: pointer; border: 1.5px solid var(--border);
            background: var(--warm-wh); color: var(--muted); transition: all 0.25s;
            display: flex; align-items: center; gap: 0.45rem;
        }
        .group-chip:hover { border-color: var(--gold); color: var(--navy); }
        .group-chip.active {
            background: linear-gradient(135deg, var(--navy), var(--navy2));
            color: var(--gold-lt); border-color: var(--navy);
        }
        .group-chip .delete-group {
            font-size: 0.78rem; margin-left: 0.2rem; opacity: 0.5; cursor: pointer;
        }
        .group-chip .delete-group:hover { opacity: 1; }
        .add-group-btn {
            padding: 0.48rem 1.1rem; border-radius: 20px;
            font-size: 0.83rem; font-weight: 600;
            cursor: pointer; border: 2px dashed var(--border);
            background: transparent; color: #A0937E;
            transition: all 0.25s; font-family: 'Inter', sans-serif;
        }
        .add-group-btn:hover { border-color: var(--gold); color: var(--gold); }

        /* ========== MODAL ========== */
        .modal-overlay {
            display: none;
            position: fixed; top: 0; left: 0; width: 100%; height: 100%;
            background: rgba(13,27,42,0.7); z-index: 500;
            align-items: center; justify-content: center;
            backdrop-filter: blur(5px);
        }
        .modal-overlay.show { display: flex; }

        .modal-box {
            background: var(--warm-wh); border-radius: 18px;
            max-width: 720px; width: 92%; max-height: 92vh;
            overflow-y: auto; box-shadow: 0 32px 96px rgba(0,0,0,0.4);
            position: relative; border-top: 4px solid var(--gold);
        }
        .modal-close {
            position: absolute; top: 14px; right: 16px;
            background: rgba(0,0,0,0.55); color: white;
            border: none; border-radius: 50%; width: 36px; height: 36px;
            font-size: 1.3rem; cursor: pointer; z-index: 510;
            display: flex; align-items: center; justify-content: center;
            transition: background 0.2s;
        }
        .modal-close:hover { background: var(--burgundy); }
        .modal-img {
            width: 100%; height: 340px;
            object-fit: cover; display: block; border-radius: 14px 14px 0 0;
        }
        .modal-body { padding: 1.5rem 2rem 2rem; }
        .modal-title {
            display: flex; justify-content: space-between; align-items: flex-start;
            margin-bottom: 1.2rem; padding-bottom: 1rem;
            border-bottom: 2px solid var(--border);
        }
        .modal-title h2 {
            font-family: 'Playfair Display', serif;
            color: var(--navy); font-size: 1.4rem;
        }
        .modal-title span { color: var(--gold); font-size: 0.88rem; font-weight: 600; }
        .modal-props {
            display: grid; grid-template-columns: 1fr 1fr 1fr;
            gap: 1.1rem; margin-bottom: 1.5rem;
        }
        @media (max-width: 500px) { .modal-props { grid-template-columns: 1fr 1fr; } }
        .modal-prop-label {
            color: #B0A898; font-weight: 700;
            text-transform: uppercase; font-size: 0.68rem; letter-spacing: 0.7px;
        }
        .modal-prop-value { color: var(--charcoal); font-weight: 600; font-size: 0.98rem; margin-top: 0.15rem; }
        .modal-wishlist-bar {
            display: flex; gap: 0.8rem; align-items: center;
            padding-top: 1rem; border-top: 1px solid var(--border);
        }
        .modal-wishlist-bar select {
            flex: 1; padding: 0.72rem 0.95rem;
            border: 1.5px solid var(--border); border-radius: 9px;
            font-size: 0.93rem; font-family: 'Inter', sans-serif;
            background: white; color: var(--charcoal);
        }
        .modal-wishlist-bar select:focus {
            outline: none; border-color: var(--gold);
            box-shadow: 0 0 0 3px rgba(196,149,42,0.14);
        }
        .btn-add-wish {
            padding: 0.72rem 1.5rem;
            background: linear-gradient(135deg, var(--burgundy), #A01A3E);
            color: white; border: none; border-radius: 9px;
            font-size: 0.88rem; font-weight: 700; cursor: pointer;
            white-space: nowrap; transition: all 0.25s;
            font-family: 'Inter', sans-serif; letter-spacing: 0.3px;
        }
        .btn-add-wish:hover { transform: translateY(-2px); box-shadow: 0 6px 18px rgba(139,21,53,0.38); }
        .modal-msg {
            margin-top: 0.7rem; font-size: 0.85rem; color: #2e7d32;
            font-weight: 600; display: none;
            background: #f1f8e9; padding: 0.5rem 0.8rem; border-radius: 6px;
        }
        .modal-msg.show { display: block; }

        /* ========== NEW GROUP MODAL ========== */
        .group-modal-overlay {
            display: none;
            position: fixed; top: 0; left: 0; width: 100%; height: 100%;
            background: rgba(13,27,42,0.65); z-index: 600;
            align-items: center; justify-content: center;
            backdrop-filter: blur(3px);
        }
        .group-modal-overlay.show { display: flex; }
        .group-modal-box {
            background: var(--warm-wh); border-radius: 16px;
            padding: 2.2rem 2rem; width: 92%; max-width: 400px;
            box-shadow: 0 20px 56px rgba(0,0,0,0.35);
            border-top: 4px solid var(--gold);
        }
        .group-modal-box h3 {
            font-family: 'Playfair Display', serif;
            color: var(--navy); margin-bottom: 1.2rem; font-size: 1.2rem;
        }
        .group-modal-box input {
            width: 100%; padding: 0.85rem 1rem;
            border: 1.5px solid var(--border); border-radius: 10px;
            font-size: 1rem; font-family: 'Inter', sans-serif;
            margin-bottom: 1.2rem; background: white;
        }
        .group-modal-box input:focus {
            outline: none; border-color: var(--gold);
            box-shadow: 0 0 0 3px rgba(196,149,42,0.14);
        }
        .group-modal-actions { display: flex; gap: 0.8rem; justify-content: flex-end; }
        .group-modal-actions button { padding: 0.62rem 1.4rem; border-radius: 8px; font-size: 0.9rem; font-weight: 600; font-family: 'Inter', sans-serif; }
        .btn-cancel-group { background: #EDE8DF; color: var(--muted); border: 1px solid var(--border); cursor: pointer; }
        .btn-cancel-group:hover { background: #E0D8CE; }
        .btn-save-group {
            background: linear-gradient(135deg, var(--burgundy), #A01A3E);
            color: white; border: none; cursor: pointer;
        }
        .btn-save-group:hover { transform: translateY(-1px); box-shadow: 0 4px 14px rgba(139,21,53,0.35); }

        /* ========== SCROLLBAR ========== */
        ::-webkit-scrollbar { width: 8px; height: 8px; }
        ::-webkit-scrollbar-track { background: transparent; }
        ::-webkit-scrollbar-thumb { background: rgba(196,149,42,0.3); border-radius: 4px; }
        ::-webkit-scrollbar-thumb:hover { background: rgba(196,149,42,0.5); }

        /* ==========================================================
           UI ENHANCEMENTS — depth, motion & polish (cascade overrides)
           ========================================================== */
        html { scroll-behavior: smooth; }
        body { -webkit-font-smoothing: antialiased; text-rendering: optimizeLegibility; }

        @keyframes fadeUp { from { opacity: 0; transform: translateY(12px); } to { opacity: 1; transform: none; } }
        @keyframes popIn  { from { opacity: 0; transform: scale(.96); }       to { opacity: 1; transform: none; } }

        /* Ambient depth behind the working area */
        .main-content {
            background:
                radial-gradient(1100px 420px at 100% -8%, rgba(196,149,42,0.10), transparent 60%),
                radial-gradient(900px 400px at -10% 112%, rgba(27,42,74,0.07), transparent 55%),
                var(--cream);
        }
        .page-section.active { animation: fadeUp .38s cubic-bezier(.2,.7,.2,1) both; }

        /* Header — a touch more presence */
        .header { box-shadow: 0 4px 22px rgba(13,27,42,0.30); }
        .header-brand-icon { box-shadow: 0 4px 14px rgba(196,149,42,0.5); }

        /* Sidebar nav — floating pill interaction */
        .sidebar { background: linear-gradient(185deg, var(--navy) 0%, #16223c 100%); }
        .nav-item { margin: .15rem .7rem; border-radius: 11px; padding-left: 1.1rem; transition: background .2s, color .2s, transform .15s; }
        .nav-item:hover { transform: translateX(2px); }
        .nav-item.active {
            background: linear-gradient(135deg, rgba(196,149,42,0.22), rgba(196,149,42,0.10));
            box-shadow: inset 0 0 0 1px rgba(196,149,42,0.25);
        }

        /* Cards — softer, deeper, gently interactive */
        .card {
            border-radius: 18px;
            box-shadow: 0 8px 26px rgba(27,42,74,0.09), 0 1px 3px rgba(0,0,0,0.04);
            transition: box-shadow .3s ease, transform .3s ease;
        }
        .card:hover { box-shadow: 0 16px 44px rgba(27,42,74,0.14); }

        /* Page headers */
        .page-header h2 { font-family: 'Playfair Display', serif; color: var(--navy); font-size: 1.55rem; }

        /* Inputs & selects — rounder, custom caret, hover affordance */
        input[type="text"], input[type="number"], select {
            border-radius: 12px; padding: 0.82rem 1rem; background: #fffdf9;
        }
        input[type="text"]:hover, input[type="number"]:hover, select:hover { border-color: #d8c79a; }
        select {
            appearance: none; -webkit-appearance: none; -moz-appearance: none;
            background-image: url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' width='12' height='8' viewBox='0 0 12 8'><path fill='%23C4952A' d='M6 8 0 2 1.4.6 6 5.2 10.6.6 12 2z'/></svg>");
            background-repeat: no-repeat; background-position: right 1rem center; padding-right: 2.5rem;
        }

        /* Buttons — shimmer sweep + crisp press */
        button { position: relative; overflow: hidden; }
        button::after {
            content: ''; position: absolute; top: 0; left: -120%; width: 100%; height: 100%;
            background: linear-gradient(120deg, transparent, rgba(255,255,255,0.28), transparent);
            transform: skewX(-20deg); transition: left .55s ease; pointer-events: none;
        }
        button:hover::after { left: 130%; }
        .btn-search { box-shadow: 0 5px 16px rgba(139,21,53,0.26); }
        .btn-search:hover { box-shadow: 0 12px 28px rgba(139,21,53,0.42); }
        .btn-search:active, .btn-reset:active, .btn-save-group:active { transform: translateY(0) scale(.98); }

        /* Results banner */
        .results-info { box-shadow: 0 4px 16px rgba(196,149,42,0.12); }

        /* Sample cards — image zoom, lift, staggered entrance */
        .samples-grid { gap: 1.6rem; }
        .sample-card { border-radius: 18px; animation: popIn .32s ease both; }
        .sample-card:hover { transform: translateY(-6px); box-shadow: 0 20px 44px rgba(27,42,74,0.18); }
        .sample-card-img { transition: transform .55s cubic-bezier(.2,.7,.2,1); }
        .sample-card:hover .sample-card-img { transform: scale(1.08); }
        .sample-card-no { font-size: 1.06rem; }

        /* Count badge */
        .count-badge {
            background: linear-gradient(135deg, var(--navy), var(--navy2));
            color: var(--gold-lt); border: 1px solid rgba(196,149,42,0.3);
            box-shadow: 0 3px 10px rgba(27,42,74,0.18);
        }

        /* Modal — deeper backdrop & entrance */
        .modal-box { border-radius: 18px; box-shadow: 0 30px 80px rgba(0,0,0,0.5); animation: popIn .28s ease both; }
    </style>
</head>
<body>
    <!-- ======= HEADER ======= -->
    <div class="header">
        <div class="header-brand">
            <div class="header-brand-icon">&#129525;</div>
            <div class="header-brand-text">
                <h1>FabricSample</h1>
                <p>Smart Search Platform</p>
            </div>
        </div>
        <div class="header-right">
            <div class="header-user">
                <div class="header-avatar">{{ username[0]|upper }}</div>
                <span class="header-username">{{ username }}</span>
            </div>
            <a href="/logout" class="btn-logout">Sign Out</a>
        </div>
    </div>

    <div class="app-layout">
        <!-- ======= SIDEBAR ======= -->
        <div class="sidebar">
            <div class="sidebar-label">Navigation</div>
            <div class="nav-item active" onclick="switchPage('search', this)">
                <span class="nav-icon">&#128269;</span>
                <span>Search</span>
            </div>
            <div class="nav-item" onclick="switchPage('wishlist', this)">
                <span class="nav-icon">&#10084;</span>
                <span>Wishlist</span>
                <span class="nav-badge" id="wishlistBadge">0</span>
            </div>
            <div class="sidebar-divider"></div>
            <div class="nav-item" onclick="switchPage('data', this)">
                <span class="nav-icon">&#128202;</span>
                <span>All Samples</span>
            </div>
            <div class="nav-item" onclick="switchPage('upload', this)">
                <span class="nav-icon">&#128228;</span>
                <span>Upload Samples</span>
            </div>
        </div>

        <!-- ======= MAIN CONTENT ======= -->
        <div class="main-content">

            <!-- ---- SEARCH PAGE ---- -->
            <div id="page-search" class="page-section active">
                <div class="container">
                    <div class="card">
                        <div class="card-title">Search Fabric Samples</div>
                        <form id="searchForm">
                            <div class="form-grid">
                                <div class="form-group">
                                    <label for="product_type">Product Type</label>
                                    <select id="product_type" name="product_type">
                                        <option value="ALL">All Products</option>
                                        <option value="DYED">DYED</option>
                                        <option value="PRINT">PRINT</option>
                                        <option value="CHECKS">CHECKS</option>
                                        <option value="STRIPES">STRIPES</option>
                                        <option value="WHITE">WHITE</option>
                                        <option value="YD+PRINT">YD+PRINT</option>
                                        <option value="WHITE+PRINT">WHITE+PRINT</option>
                                        <option value="DYED+PRINT">DYED+PRINT</option>
                                        <option value="YD+PIGMENT PRINT">YD+PIGMENT PRINT</option>
                                    </select>
                                </div>
                                <div class="form-group">
                                    <label for="weave">Weave</label>
                                    <select id="weave" name="weave">
                                        <option value="ALL">All Weaves</option>
                                        <option value="PLAIN">PLAIN</option>
                                        <option value="TWILL">TWILL</option>
                                        <option value="DOBBY">DOBBY</option>
                                        <option value="SATIN">SATIN</option>
                                    </select>
                                </div>
                                <div class="form-group">
                                    <label for="yarn">Yarn Type</label>
                                    <select id="yarn" name="yarn">
                                        <option value="ALL">All Yarn Types</option>
                                        <option value="COMPACT">COMPACT</option>
                                        <option value="COMBED">COMBED</option>
                                        <option value="SLUB">SLUB</option>
                                        <option value="TFO">TFO</option>
                                        <option value="CARDED">CARDED</option>
                                    </select>
                                </div>
                                <div class="form-group">
                                    <label for="blend">Composition / Blend</label>
                                    <input type="text" id="blend" name="blend" placeholder="e.g., cotton, modal, viscose">
                                </div>
                                <div class="form-group">
                                    <label for="gsm_min">GSM Min</label>
                                    <input type="number" id="gsm_min" name="gsm_min" placeholder="e.g., 100" min="0">
                                </div>
                                <div class="form-group">
                                    <label for="gsm_max">GSM Max</label>
                                    <input type="number" id="gsm_max" name="gsm_max" placeholder="e.g., 200" min="0">
                                </div>
                                <div class="form-group full-width">
                                    <label for="feel_terms">Performance / Feel Terms</label>
                                    <input type="text" id="feel_terms" name="feel_terms" placeholder="e.g., soft feel, shiny, stretchable, crisp">
                                </div>
                            </div>
                            <div class="button-group">
                                <button type="reset" class="btn-reset">&#10006; Clear</button>
                                <button type="submit" class="btn-search">&#128269; Search Samples</button>
                            </div>
                        </form>
                    </div>

                    <div class="loading" id="loading">
                        <div class="spinner"></div><p>Searching samples...</p>
                    </div>
                    <div class="error-message" id="errorMessage"></div>

                    <div class="results-section" id="resultsSection">
                        <div class="results-info">
                            <h3><span id="resultCount">0</span> samples found</h3>
                            <div id="termsDisplay"></div>
                        </div>
                        <div class="samples-grid" id="resultsContent"></div>
                    </div>
                </div>
            </div>

            <!-- ---- WISHLIST PAGE ---- -->
            <div id="page-wishlist" class="page-section">
                <div class="container">
                    <div class="page-header">
                        <h2>&#10084; My Wishlist</h2>
                        <span class="count-badge" id="wishlistCount">0 samples</span>
                    </div>
                    <div class="group-bar" id="groupBar">
                        <!-- group chips rendered by JS -->
                    </div>
                    <div id="wishlistContent">
                        <div class="empty-state">
                            <div class="empty-icon">&#10084;</div>
                            <p>Your wishlist is empty</p>
                            <p class="sub">Create a group, then click on any sample to add it</p>
                        </div>
                    </div>
                </div>
            </div>

            <!-- ---- DATA PAGE ---- -->
            <div id="page-data" class="page-section">
                <div class="container">
                    <div class="page-header">
                        <h2>&#128202; All Fabric Samples</h2>
                        <span class="count-badge" id="dataCount">Loading...</span>
                    </div>
                    <div class="loading show" id="dataLoading">
                        <div class="spinner"></div><p>Loading all samples...</p>
                    </div>
                    <div class="samples-grid" id="dataGrid"></div>
                </div>
            </div>

            <!-- ---- UPLOAD PAGE ---- -->
            <div id="page-upload" class="page-section">
                <div class="container">
                    <div class="page-header">
                        <h2>&#128228; Upload Samples</h2>
                    </div>
                    <div class="card" id="uploadCard">
                        <div class="card-title">Add samples from an Excel sheet</div>
                        <p style="color:var(--muted);font-size:0.92rem;margin-bottom:1.4rem;">
                            Drop an <strong>.xlsx</strong> file in the standard layout (attributes down the left
                            column &mdash; ARTICLE, SAMPLE NO., IMAGE, PRODUCT, YARN, COUNT, CONSTRUCTION, BLEND,
                            WEAVE, FINISH, GSM &mdash; one sample per column, images in the IMAGE row). You'll see a
                            full preview to review <em>before</em> anything is saved.
                        </p>
                        <div id="dropZone" style="border:2px dashed var(--gold);border-radius:12px;padding:2.4rem;text-align:center;background:#fffaf0;cursor:pointer;transition:all .2s;">
                            <div style="font-size:2.4rem;margin-bottom:.6rem;">&#128193;</div>
                            <div id="dropLabel" style="font-weight:600;color:var(--navy);">Click to choose an Excel file, or drag &amp; drop it here</div>
                            <div id="fileName" style="margin-top:.5rem;color:var(--muted);font-size:0.88rem;"></div>
                            <input type="file" id="uploadFile" accept=".xlsx,.xlsm" style="display:none;">
                        </div>
                    </div>

                    <div class="loading" id="uploadLoading">
                        <div class="spinner"></div><p id="uploadLoadingText">Reading the sheet &amp; building preview...</p>
                    </div>
                    <div class="error-message" id="uploadError"></div>
                    <div id="uploadPreview"></div>
                    <div id="uploadResult"></div>
                </div>
            </div>

        </div>
    </div>

    <!-- ======= SAMPLE DETAIL MODAL ======= -->
    <div class="modal-overlay" id="sampleModal">
        <div class="modal-box">
            <button class="modal-close" onclick="closeModal()">&times;</button>
            <img class="modal-img" id="modalImg" src="" alt="">
            <div class="modal-body">
                <div class="modal-title">
                    <h2 id="modalSampleNo"></h2>
                    <span id="modalArticle"></span>
                </div>
                <div class="modal-props" id="modalProps"></div>
                <div class="modal-wishlist-bar">
                    <select id="modalGroupSelect">
                        <option value="">— Select Wishlist Group —</option>
                    </select>
                    <button class="btn-add-wish" onclick="addFromModal()">&#10084; Add to Wishlist</button>
                </div>
                <div class="modal-msg" id="modalMsg">&#10003; Added to wishlist!</div>
            </div>
        </div>
    </div>

    <!-- ======= NEW GROUP MODAL ======= -->
    <div class="group-modal-overlay" id="groupModal">
        <div class="group-modal-box">
            <h3>Create Wishlist Group</h3>
            <input type="text" id="groupNameInput" placeholder="e.g., Arvind, Aditya Birla...">
            <div class="group-modal-actions">
                <button class="btn-cancel-group" onclick="closeGroupModal()">Cancel</button>
                <button class="btn-save-group" onclick="saveNewGroup()">&#10003; Create</button>
            </div>
        </div>
    </div>

    <script>
        let allGroups = [];       // [{id, name}]
        let wishlistData = null;  // full wishlist response
        let activeGroupId = null; // currently viewed group in wishlist tab
        let allSamplesCache = []; // for detail modal

        /* ---- PAGE SWITCHING ---- */
        function switchPage(page, el) {
            document.querySelectorAll('.nav-item').forEach(n => n.classList.remove('active'));
            document.querySelectorAll('.page-section').forEach(p => p.classList.remove('active'));
            document.getElementById('page-' + page).classList.add('active');
            if (el) el.classList.add('active');
            if (page === 'data' && !window._dataLoaded) loadAllSamples();
            if (page === 'wishlist') loadWishlist();
        }

        /* ---- UPLOAD SAMPLES (preview -> confirm) ---- */
        let _uploadFileObj = null;
        function _initUpload() {
            const zone = document.getElementById('dropZone');
            const input = document.getElementById('uploadFile');
            if (!zone || !input) return;
            zone.addEventListener('click', () => input.click());
            input.addEventListener('change', () => { if (input.files.length) setUploadFile(input.files[0]); });
            zone.addEventListener('dragover', (e) => { e.preventDefault(); zone.style.background = '#fff2d6'; zone.style.borderColor = '#8B1535'; });
            zone.addEventListener('dragleave', () => { zone.style.background = '#fffaf0'; zone.style.borderColor = 'var(--gold)'; });
            zone.addEventListener('drop', (e) => {
                e.preventDefault(); zone.style.background = '#fffaf0'; zone.style.borderColor = 'var(--gold)';
                if (e.dataTransfer.files.length) setUploadFile(e.dataTransfer.files[0]);
            });
        }
        function setUploadFile(file) {
            _uploadFileObj = file;
            document.getElementById('fileName').textContent = '✓ ' + file.name;
            document.getElementById('uploadError').classList.remove('show');
            document.getElementById('uploadResult').innerHTML = '';
            previewUpload();   // auto-preview on select
        }
        function clearUpload() {
            _uploadFileObj = null;
            document.getElementById('uploadFile').value = '';
            document.getElementById('fileName').textContent = '';
            document.getElementById('uploadError').classList.remove('show');
            document.getElementById('uploadPreview').innerHTML = '';
            document.getElementById('uploadResult').innerHTML = '';
        }
        function _esc(v) {
            return (v === null || v === undefined ? '' : String(v))
                .replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;');
        }
        function _setLoading(on, text) {
            const l = document.getElementById('uploadLoading');
            if (text) document.getElementById('uploadLoadingText').textContent = text;
            l.classList.toggle('show', !!on);
        }

        async function previewUpload() {
            if (!_uploadFileObj) return;
            const err = document.getElementById('uploadError');
            const preview = document.getElementById('uploadPreview');
            err.classList.remove('show'); preview.innerHTML = '';
            _setLoading(true, 'Reading the sheet & building preview...');
            try {
                const fd = new FormData(); fd.append('file', _uploadFileObj);
                const resp = await fetch('/api/upload-samples/preview', { method: 'POST', body: fd });
                const data = await resp.json();
                _setLoading(false);
                if (!resp.ok) { err.textContent = data.error || 'Could not read the file.'; err.classList.add('show'); return; }
                renderPreview(data);
            } catch (e) {
                _setLoading(false);
                err.textContent = 'Preview failed: ' + e; err.classList.add('show');
            }
        }

        function _chip(label, value, color) {
            return '<div style="background:' + color + '1a;border:1px solid ' + color + '55;color:' + color +
                ';border-radius:10px;padding:.5rem .9rem;font-weight:700;font-size:.82rem;display:flex;flex-direction:column;align-items:center;min-width:78px;">' +
                '<span style="font-size:1.25rem;line-height:1;">' + value + '</span>' +
                '<span style="font-weight:600;opacity:.8;font-size:.68rem;text-transform:uppercase;letter-spacing:.4px;margin-top:.2rem;">' + label + '</span></div>';
        }

        function renderPreview(data) {
            const s = data.summary;
            const cols = ['product', 'yarn', 'count', 'construction', 'blend', 'weave', 'finish', 'gsm'];
            let h = '<div class="card" style="border-top-color:var(--navy);">';
            h += '<div class="card-title">Preview &mdash; review before saving</div>';

            // summary chips
            h += '<div style="display:flex;gap:.7rem;flex-wrap:wrap;margin-bottom:1.2rem;">';
            h += _chip('Total', s.total, '#1B2A4A');
            h += _chip('New', s.new, '#2e7d32');
            h += _chip('Updates', s.updates, '#C4952A');
            h += _chip('Images', s.with_image, '#0097a7');
            if (s.without_image) h += _chip('No image', s.without_image, '#8B1535');
            if (s.with_issues) h += _chip('Need check', s.with_issues, '#8B1535');
            h += '</div>';

            if (s.with_issues) {
                h += '<div style="background:#fff8e1;border:1px solid #ffe0a3;color:#8a5a00;border-radius:8px;padding:.6rem .9rem;font-size:.85rem;margin-bottom:1rem;">' +
                    '&#9888; ' + s.with_issues + ' sample(s) have highlighted fields (missing or unreadable). You can still save &mdash; flagged numbers default to 0.</div>';
            }

            // table
            h += '<div style="overflow-x:auto;border:1px solid var(--border);border-radius:10px;">';
            h += '<table style="border-collapse:collapse;width:100%;font-size:.82rem;white-space:nowrap;">';
            h += '<thead><tr style="background:var(--navy);color:#fff;">';
            ['', 'Sample', 'Status', 'Article', 'Product', 'Yarn', 'Count', 'Constr.', 'Blend', 'Weave', 'Finish', 'GSM']
                .forEach(t => { h += '<th style="padding:.55rem .7rem;text-align:left;font-weight:600;">' + t + '</th>'; });
            h += '</tr></thead><tbody>';

            data.records.forEach((r, i) => {
                const bg = i % 2 ? '#fbf7ee' : '#fff';
                const iss = new Set(r.issues || []);
                h += '<tr style="background:' + bg + ';border-top:1px solid var(--border);">';
                // thumbnail
                h += '<td style="padding:.35rem .5rem;">';
                if (r.thumb) h += '<img src="' + r.thumb + '" style="width:42px;height:42px;object-fit:cover;border-radius:6px;border:1px solid var(--border);">';
                else h += '<span title="no image" style="display:inline-flex;width:42px;height:42px;border-radius:6px;background:#f0e9da;border:1px dashed #c9bfa6;align-items:center;justify-content:center;color:#b26a00;font-size:1rem;">&#128247;</span>';
                h += '</td>';
                // sample no
                h += '<td style="padding:.45rem .7rem;font-weight:700;color:var(--navy);">' + _esc(r.sample_no) +
                    (iss.has('duplicate') ? ' <span title="duplicated in sheet" style="color:#8B1535;">&#9888;</span>' : '') + '</td>';
                // status badge
                const isNew = r.status === 'new';
                h += '<td style="padding:.45rem .7rem;"><span style="font-size:.66rem;font-weight:700;text-transform:uppercase;letter-spacing:.4px;padding:.16rem .5rem;border-radius:20px;background:' +
                    (isNew ? '#e6f4ea;color:#2e7d32;' : '#fbf0d5;color:#946a12;') + '">' + (isNew ? 'New' : 'Update') + '</span></td>';
                // article
                const aBad = iss.has('article');
                h += '<td style="padding:.45rem .7rem;' + (aBad ? 'background:#ffe3e3;color:#8B1535;font-style:italic;' : '') + '">' + (_esc(r.article) || (aBad ? 'missing' : '')) + '</td>';
                // other cols
                cols.forEach(c => {
                    const bad = iss.has(c) || (c === 'gsm' && iss.has('gsm')) || (c === 'count' && iss.has('count'));
                    let val = r[c];
                    if (c === 'gsm' && (val === 0 || val === null)) val = bad ? '0' : val;
                    h += '<td style="padding:.45rem .7rem;' + (bad ? 'background:#ffe3e3;color:#8B1535;font-weight:600;' : '') + '">' + _esc(val) + '</td>';
                });
                h += '</tr>';
            });
            h += '</tbody></table></div>';

            if (data.warnings && data.warnings.length) {
                h += '<div style="margin-top:1rem;color:#b26a00;font-size:0.82rem;">';
                h += '<strong>Notes:</strong><ul style="margin:.4rem 0 0 1.2rem;">';
                data.warnings.forEach(w => { h += '<li>' + _esc(w) + '</li>'; });
                h += '</ul></div>';
            }

            // action buttons
            h += '<div class="button-group">';
            h += '<button type="button" class="btn-reset" onclick="clearUpload()">&#10006; Cancel</button>';
            h += '<button type="button" class="btn-search" id="confirmUploadBtn" onclick="confirmUpload()">&#10003; Confirm &amp; Save ' + s.total + ' Sample(s)</button>';
            h += '</div></div>';

            document.getElementById('uploadPreview').innerHTML = h;
        }

        async function confirmUpload() {
            if (!_uploadFileObj) return;
            const err = document.getElementById('uploadError');
            const result = document.getElementById('uploadResult');
            const btn = document.getElementById('confirmUploadBtn');
            err.classList.remove('show'); result.innerHTML = '';
            if (btn) btn.disabled = true;
            _setLoading(true, 'Saving data & uploading images...');
            try {
                const fd = new FormData(); fd.append('file', _uploadFileObj);
                const resp = await fetch('/api/upload-samples', { method: 'POST', body: fd });
                const data = await resp.json();
                _setLoading(false);
                if (!resp.ok) { err.textContent = data.error || 'Upload failed.'; err.classList.add('show'); if (btn) btn.disabled = false; return; }
                let html = '<div class="card" style="border-top-color:#2e7d32;">';
                html += '<div class="card-title" style="color:#2e7d32;">&#10003; Saved successfully</div>';
                html += '<div class="modal-props">';
                html += prop('Samples added / updated', data.samples_added);
                html += prop('Sample range', data.sample_range[0] + ' – ' + data.sample_range[1]);
                html += prop('Images uploaded', data.images_uploaded);
                if (data.images_failed) html += prop('Images failed', data.images_failed);
                if (data.no_image && data.no_image.length) html += prop('Samples without an image', data.no_image.join(', '));
                html += '</div></div>';
                result.innerHTML = html;
                document.getElementById('uploadPreview').innerHTML = '';
                clearUpload();
                window._dataLoaded = false;  // force "All Samples" to refresh next visit
                result.scrollIntoView({ behavior: 'smooth', block: 'center' });
            } catch (e) {
                _setLoading(false);
                err.textContent = 'Upload failed: ' + e; err.classList.add('show');
                if (btn) btn.disabled = false;
            }
        }
        _initUpload();

        /* ---- CARD BUILDER ---- */
        function buildSampleCard(sample, rankInfo) {
            let html = '<div class="sample-card" id="card-' + sample.sample_no + '" onclick="openModal(' + sample.sample_no + ')">';
            if (rankInfo) {
                html += '<span class="rank-badge ' + rankInfo.cls + '">#' + rankInfo.rank + '</span>';
                html += '<span class="rank-label">' + rankInfo.label + '</span>';
            }
            html += '<img class="sample-card-img" src="/sample-image/' + sample.sample_no + '" alt="Sample ' + sample.sample_no + '" onerror="this.style.background=\'#ddd\';this.style.height=\'140px\'">';
            html += '<div class="sample-card-body">';
            html += '<div class="sample-card-header">';
            html += '<span class="sample-card-no">' + sample.sample_no + '</span>';
            html += '<span class="sample-card-article">' + sample.article + '</span>';
            html += '</div>';
            html += '<div class="sample-card-props">';
            html += prop('Product', sample.product);
            html += prop('Yarn', sample.yarn);
            html += prop('Count', sample.count);
            html += prop('GSM', sample.gsm);
            html += prop('Blend', sample.blend);
            html += prop('Weave', sample.weave);
            html += prop('Finish', sample.finish);
            html += '</div></div></div>';
            return html;
        }

        function prop(label, value) {
            return '<div class="sample-prop"><div class="sample-prop-label">' + label + '</div><div class="sample-prop-value">' + value + '</div></div>';
        }

        /* ============ SAMPLE DETAIL MODAL ============ */
        function openModal(sampleNo) {
            const sample = findSample(sampleNo);
            if (!sample) return;
            document.getElementById('modalImg').src = '/sample-image/' + sample.sample_no;
            document.getElementById('modalSampleNo').textContent = 'Sample ' + sample.sample_no;
            document.getElementById('modalArticle').textContent = sample.article;
            const props = [
                ['Product', sample.product], ['Yarn', sample.yarn], ['Count', sample.count],
                ['Construction', sample.construction || '-'], ['Blend', sample.blend], ['Weave', sample.weave],
                ['Finish', sample.finish], ['GSM', sample.gsm], ['Count Avg', sample.count_avg || '-']
            ];
            let ph = '';
            props.forEach(function(p) {
                ph += '<div class="modal-prop"><div class="modal-prop-label">' + p[0] + '</div><div class="modal-prop-value">' + p[1] + '</div></div>';
            });
            document.getElementById('modalProps').innerHTML = ph;
            // populate group dropdown
            refreshGroupDropdown();
            document.getElementById('modalMsg').classList.remove('show');
            document.getElementById('sampleModal').dataset.sampleNo = sampleNo;
            document.getElementById('sampleModal').classList.add('show');
        }

        function closeModal() {
            document.getElementById('sampleModal').classList.remove('show');
        }

        function refreshGroupDropdown() {
            const sel = document.getElementById('modalGroupSelect');
            sel.innerHTML = '<option value="">-- Select Wishlist Group --</option>';
            allGroups.forEach(function(g) {
                sel.innerHTML += '<option value="' + g.id + '">' + g.name + '</option>';
            });
        }

        async function addFromModal() {
            const groupId = document.getElementById('modalGroupSelect').value;
            const sampleNo = document.getElementById('sampleModal').dataset.sampleNo;
            if (!groupId) { alert('Please select a wishlist group first'); return; }
            await fetch('/api/wishlist/add', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({sample_no: parseInt(sampleNo), group_id: parseInt(groupId)})
            });
            const msg = document.getElementById('modalMsg');
            const groupName = allGroups.find(g => g.id == groupId).name;
            msg.textContent = 'Added to "' + groupName + '" wishlist!';
            msg.classList.add('show');
            setTimeout(function(){ msg.classList.remove('show'); }, 2000);
            loadWishlistBadge();
        }

        // click outside modal to close
        document.getElementById('sampleModal').addEventListener('click', function(e) {
            if (e.target === this) closeModal();
        });

        function findSample(sampleNo) {
            if (allSamplesCache.length > 0) {
                const found = allSamplesCache.find(s => s.sample_no == sampleNo);
                if (found) return found;
            }
            // try search results
            const cards = document.querySelectorAll('.sample-card');
            // fallback: fetch sync is not ideal, but we cache on first data/search load
            return null;
        }

        /* ============ WISHLIST GROUPS ============ */
        function openGroupModal() {
            document.getElementById('groupNameInput').value = '';
            document.getElementById('groupModal').classList.add('show');
        }
        function closeGroupModal() {
            document.getElementById('groupModal').classList.remove('show');
        }

        async function saveNewGroup() {
            const name = document.getElementById('groupNameInput').value.trim();
            if (!name) return;
            const res = await fetch('/api/wishlist/groups/create', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({name: name})
            });
            const data = await res.json();
            if (data.error) { alert(data.error); return; }
            closeGroupModal();
            await loadGroupsList();
            loadWishlist();
        }

        async function deleteGroup(groupId, e) {
            e.stopPropagation();
            if (!confirm('Delete this group and all its items?')) return;
            await fetch('/api/wishlist/groups/delete', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({group_id: groupId})
            });
            await loadGroupsList();
            activeGroupId = null;
            loadWishlist();
        }

        async function loadGroupsList() {
            const res = await fetch('/api/wishlist/groups');
            const data = await res.json();
            allGroups = data.groups;
        }

        function renderGroupBar(groups, activeId) {
            const bar = document.getElementById('groupBar');
            let html = '';
            groups.forEach(function(g) {
                const cls = g.id === activeId ? 'group-chip active' : 'group-chip';
                html += '<div class="' + cls + '" onclick="selectGroup(' + g.id + ')">';
                html += '<span>' + g.name + '</span>';
                html += '<span class="delete-group" onclick="deleteGroup(' + g.id + ', event)" title="Delete group">&times;</span>';
                html += '</div>';
            });
            html += '<button class="add-group-btn" onclick="openGroupModal()">+ New Group</button>';
            bar.innerHTML = html;
        }

        function selectGroup(groupId) {
            activeGroupId = groupId;
            renderWishlistContent();
        }

        async function loadWishlist() {
            await loadGroupsList();
            const res = await fetch('/api/wishlist');
            wishlistData = await res.json();
            // cache samples
            wishlistData.groups.forEach(function(g) {
                g.samples.forEach(function(s) {
                    if (!allSamplesCache.find(x => x.sample_no === s.sample_no)) allSamplesCache.push(s);
                });
            });
            updateWishlistBadge();
            document.getElementById('wishlistCount').textContent = wishlistData.total_count + ' samples';
            if (allGroups.length > 0 && !activeGroupId) activeGroupId = allGroups[0].id;
            renderGroupBar(allGroups, activeGroupId);
            renderWishlistContent();
        }

        function renderWishlistContent() {
            renderGroupBar(allGroups, activeGroupId);
            const container = document.getElementById('wishlistContent');
            if (!wishlistData || allGroups.length === 0) {
                container.innerHTML = '<div class="empty-state"><div class="empty-icon">&#10084;</div><p>No groups yet</p><p class="sub">Click "+ New Group" to create a wishlist group</p></div>';
                return;
            }
            const grp = wishlistData.groups.find(g => g.group_id === activeGroupId);
            if (!grp || grp.samples.length === 0) {
                container.innerHTML = '<div class="empty-state"><div class="empty-icon">&#128203;</div><p>This group is empty</p><p class="sub">Click on any sample card to add it here</p></div>';
                return;
            }
            let html = '<div class="samples-grid">';
            grp.samples.forEach(function(s) {
                html += buildSampleCard(s, null);
                // add remove button overlay
                html = html.slice(0, -6); // remove last </div>
                html += '<div style="padding:0.5rem 1.1rem 1rem;text-align:right;">';
                html += '<button style="padding:0.4rem 1rem;border-radius:6px;border:1.5px solid #8B1535;background:white;color:#8B1535;font-weight:600;font-size:0.78rem;cursor:pointer;font-family:Inter,sans-serif;letter-spacing:0.2px;" onclick="event.stopPropagation();removeFromGroup(' + s.sample_no + ',' + activeGroupId + ')">&#10006; Remove</button>';
                html += '</div></div>';
            });
            html += '</div>';
            container.innerHTML = html;
        }

        async function removeFromGroup(sampleNo, groupId) {
            await fetch('/api/wishlist/remove', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({sample_no: sampleNo, group_id: groupId})
            });
            loadWishlist();
        }

        function updateWishlistBadge() {
            const cnt = wishlistData ? wishlistData.total_count : 0;
            document.getElementById('wishlistBadge').textContent = cnt;
        }

        async function loadWishlistBadge() {
            const res = await fetch('/api/wishlist');
            wishlistData = await res.json();
            updateWishlistBadge();
        }
        loadWishlistBadge();
        loadGroupsList();

        /* ============ SEARCH ============ */
        document.getElementById("searchForm").addEventListener("submit", async function(e) {
            e.preventDefault();
            const loading = document.getElementById("loading");
            const resultsSection = document.getElementById("resultsSection");
            const errorMessage = document.getElementById("errorMessage");
            loading.classList.add("show");
            resultsSection.classList.remove("show");
            errorMessage.classList.remove("show");
            try {
                const response = await fetch("/search", { method: "POST", body: new FormData(this) });
                const data = await response.json();
                loading.classList.remove("show");
                // cache for modal
                data.results.forEach(function(s) {
                    if (!allSamplesCache.find(x => x.sample_no === s.sample_no)) allSamplesCache.push(s);
                });
                displayResults(data);
                resultsSection.classList.add("show");
            } catch (error) {
                loading.classList.remove("show");
                errorMessage.textContent = "Error performing search. Please try again.";
                errorMessage.classList.add("show");
            }
        });

        function displayResults(data) {
            document.getElementById("resultCount").textContent = data.total_count;
            const termsDisplay = document.getElementById("termsDisplay");
            const resultsContent = document.getElementById("resultsContent");
            if (data.standard_terms.length > 0) {
                const tags = data.standard_terms.map(t => '<span class="tag">' + t + '</span>').join("");
                termsDisplay.innerHTML = '<p style="margin-top:0.5rem;font-weight:600;color:#1B2A4A;font-size:0.85rem;letter-spacing:0.3px;">Detected Properties:</p><div class="tags">' + tags + '</div>' +
                    '<p style="margin-top:0.6rem;font-size:0.83rem;color:#7A6E5E;">&#11088; Sorted by best match — top recommendations first</p>';
            } else { termsDisplay.innerHTML = ""; }
            if (data.results.length === 0) {
                resultsContent.innerHTML = '<div class="no-results" style="grid-column:1/-1;"><p>No samples matched your requirements.</p></div>';
                return;
            }
            const hasRank = data.standard_terms.length > 0;
            let html = '';
            data.results.forEach(function(sample, index) {
                let rankInfo = null;
                if (hasRank) {
                    const rank = sample.rank || (index + 1);
                    let cls = 'rank-normal', label = 'Match';
                    if (rank === 1) { cls = 'rank-top'; label = 'Best Match'; }
                    else if (rank <= 3) { cls = 'rank-top'; label = 'Top Pick'; }
                    else if (rank <= Math.ceil(data.total_count * 0.4)) { cls = 'rank-mid'; label = 'Good Match'; }
                    rankInfo = { rank: rank, cls: cls, label: label };
                }
                html += buildSampleCard(sample, rankInfo);
            });
            resultsContent.innerHTML = html;
        }

        /* ============ DATA PAGE ============ */
        async function loadAllSamples() {
            try {
                const response = await fetch("/api/samples");
                const data = await response.json();
                document.getElementById("dataLoading").classList.remove("show");
                document.getElementById("dataCount").textContent = data.total_count + " samples";
                allSamplesCache = data.samples;
                let html = '';
                data.samples.forEach(function(sample) { html += buildSampleCard(sample, null); });
                document.getElementById("dataGrid").innerHTML = html;
                window._dataLoaded = true;
            } catch (error) {
                document.getElementById("dataLoading").innerHTML = '<p style="color:#c62828;">Failed to load samples.</p>';
            }
        }
    </script>
</body>
</html>
"""

seed_database()

if __name__ == "__main__":
    app.run(debug=True, port=5000)
